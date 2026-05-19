#!/usr/bin/env python3
"""
Dry-run analyzer for backfilling Lori Donahue's 137 historical XLSX show files
into Supabase (shows / customers / profiles / contracts / show_deal_overrides).

What this script does NOT do:
  - Touch any Supabase tables (read-only operation)
  - Write to the migration files
  - Make any decisions about how to handle mismatches

What it DOES do:
  - Walk every XLSX in the source folder (2024 + 2025 + 2026)
  - Extract show metadata + every deal row
  - Pull current shows / customers / profiles / inventory_products from Supabase
  - Match XLSX records against existing rows and emit a JSON report

Usage:
    python3 scripts/backfill_dry_run.py \\
        "/Users/williedowns/Documents/Salta/Show Spreadsheets from Lori Donahue"

Outputs:
    /tmp/backfill-dry-run-report.json
    /tmp/backfill-dry-run-report.md  (human-readable summary)
"""

import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from typing import Any

try:
    import openpyxl
    from supabase import create_client, Client  # type: ignore
except ImportError as e:
    print(f"ERROR: missing dependency ({e}). Run: pip3 install openpyxl supabase",
          file=sys.stderr)
    sys.exit(1)

# Reuse the rep-name normalization map from the leaderboard parser
NAME_NORM = {
    "dan cesaro": "Daniel Cesaro",
    "daniel cesaro": "Daniel Cesaro",
    "gio gonzalez": "Giovanni Gonzalez",
    "giovanni gonzalez": "Giovanni Gonzalez",
    "josh wyscarver": "Josh Wyscarver",
    "joshua wyscarver": "Josh Wyscarver",
    "m. conner brady": "M. Connor Brady",
    "m. connor brady": "M. Connor Brady",
    "blake carmen": "Blake Carman",
    "blake carman": "Blake Carman",
    "alex": "Alex Broyles",
    "alex broyles": "Alex Broyles",
    "tom devlin": "Thomas Devlin",
    "thomas devlin": "Thomas Devlin",
}

def norm_rep(name):
    if not name:
        return None
    s = str(name).strip()
    return NAME_NORM.get(s.lower(), s)


SHOW_FILE_RE = re.compile(r"^(\d{2})-(\d{2})-(\d{2})\s+(.+?)\s+(?:Boat\s+|Tent\s+)?Sales\.xlsx$")

def parse_show_filename(filename):
    """Return (iso_date, city, raw_city_state) or (None, None, None)."""
    m = SHOW_FILE_RE.match(filename)
    if not m:
        return None, None, None
    mm, dd, yy = m.group(1), m.group(2), m.group(3)
    city_state = m.group(4).strip()
    iso = f"20{yy}-{mm}-{dd}"
    # city_state is like "Williamson County, TX" or "Lake Charles, LA"
    city = city_state.split(",")[0].strip()
    return iso, city, city_state


def extract_xlsx(path):
    """Return dict with show meta + list of deal records."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Sales" not in wb.sheetnames:
        return None

    ws = wb["Sales"]
    headers = [c.value for c in ws[1]]

    def col(letter):
        # Convert column letter to 0-based index
        n = 0
        for c in letter:
            n = n * 26 + (ord(c) - ord("A") + 1)
        return n - 1

    deals = []
    for row_idx in range(4, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=i + 1).value for i in range(ws.max_column)]
        status = row[col("B")]
        if not status:
            continue
        status = str(status).strip()
        # Collect up to 4 salesmen
        sms = []
        for letter in ("J", "K", "L", "M"):
            v = row[col(letter)]
            if v:
                normed = norm_rep(v)
                if normed:
                    sms.append(normed)
        if not sms:
            continue
        deal = {
            "status": status,
            "salesmen": sms,
            "model": row[col("N")],
            "sale_price": row[col("AO")],
            "total_cost": row[col("AN")],
            "last_name": row[col("D")],
            "first_name": row[col("E")],
            "city": row[col("G")],
            "state": row[col("H")],
            "zip": row[col("I")],
            "address": row[col("F")],
        }
        deals.append(deal)

    return {"deals": deals}


def get_supabase() -> Client:
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_KEY")
    if not url or not key:
        print(
            "ERROR: set NEXT_PUBLIC_SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY env vars.\n"
            "Source these from atlas-platform/.env.local",
            file=sys.stderr,
        )
        sys.exit(2)
    return create_client(url, key)


def normalize_str(s):
    if not s:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def main():
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        sys.exit(1)
    folder = sys.argv[1]
    if not os.path.isdir(folder):
        print(f"ERROR: folder not found: {folder}", file=sys.stderr)
        sys.exit(3)

    sb = get_supabase()

    # ── Load existing Supabase reference data ────────────────────────────────
    print("Fetching existing Supabase data…", file=sys.stderr)
    shows_db = sb.table("shows").select("id, name, venue_name, city, state, start_date, end_date").execute().data or []
    profiles_db = sb.table("profiles").select("id, full_name, role").execute().data or []
    products_db = (
        sb.table("products")
        .select("id, model_code, sku, name, line")
        .execute()
        .data
        or []
    )
    print(
        f"  shows in DB: {len(shows_db)}",
        f"  profiles in DB: {len(profiles_db)}",
        f"  products in DB: {len(products_db)}",
        sep="\n",
        file=sys.stderr,
    )

    # Lookup maps for matching
    show_by_date_city = {}
    for s in shows_db:
        key = (s.get("start_date"), normalize_str(s.get("city")))
        show_by_date_city.setdefault(key, []).append(s)

    profile_by_name = {}
    for p in profiles_db:
        if p.get("full_name"):
            profile_by_name[normalize_str(p["full_name"])] = p

    product_by_code = {}
    product_by_name = {}
    for prod in products_db:
        if prod.get("model_code"):
            product_by_code[normalize_str(prod["model_code"])] = prod
        if prod.get("sku"):
            product_by_code[normalize_str(prod["sku"])] = prod
        if prod.get("name"):
            product_by_name[normalize_str(prod["name"])] = prod

    # ── Walk all XLSX files ──────────────────────────────────────────────────
    file_paths = []
    for root, _, files in os.walk(folder):
        for fn in files:
            if fn.endswith(".xlsx") and not fn.startswith("~"):
                file_paths.append(os.path.join(root, fn))
    file_paths.sort(key=lambda p: os.path.basename(p))

    report = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "source_folder": folder,
        "xlsx_files_total": len(file_paths),
        "shows_in_xlsx": [],
        "shows_matched": 0,
        "shows_unmatched": 0,
        "shows_ambiguous": 0,
        "deals_total": 0,
        "deals_by_status": Counter(),
        "reps_in_xlsx": Counter(),
        "reps_matched": [],
        "reps_unmatched": [],
        "models_in_xlsx": Counter(),
        "models_matched": [],
        "models_unmatched": [],
        "unparseable_files": [],
    }

    for path in file_paths:
        fn = os.path.basename(path)
        iso, city, city_state = parse_show_filename(fn)
        if not iso:
            report["unparseable_files"].append(fn)
            continue

        try:
            data = extract_xlsx(path)
        except Exception as e:
            report["unparseable_files"].append(f"{fn} → {e}")
            continue
        if not data:
            report["unparseable_files"].append(f"{fn} → no Sales sheet")
            continue

        match_key = (iso, normalize_str(city))
        matches = show_by_date_city.get(match_key, [])
        match_status = (
            "matched" if len(matches) == 1
            else "ambiguous" if len(matches) > 1
            else "unmatched"
        )
        if match_status == "matched":
            report["shows_matched"] += 1
        elif match_status == "ambiguous":
            report["shows_ambiguous"] += 1
        else:
            report["shows_unmatched"] += 1

        show_entry = {
            "file": fn,
            "date": iso,
            "city": city_state,
            "deal_count": len(data["deals"]),
            "match_status": match_status,
            "matched_show_ids": [s["id"] for s in matches],
        }
        report["shows_in_xlsx"].append(show_entry)

        for deal in data["deals"]:
            report["deals_total"] += 1
            report["deals_by_status"][deal["status"]] += 1
            for sm in deal["salesmen"]:
                report["reps_in_xlsx"][sm] += 1
            if deal.get("model"):
                report["models_in_xlsx"][str(deal["model"]).strip()] += 1

    # Classify reps + models
    for rep, count in report["reps_in_xlsx"].most_common():
        matched_profile = profile_by_name.get(normalize_str(rep))
        bucket = "reps_matched" if matched_profile else "reps_unmatched"
        report[bucket].append({
            "name": rep,
            "deal_count": count,
            "profile_id": matched_profile["id"] if matched_profile else None,
            "profile_role": matched_profile["role"] if matched_profile else None,
        })

    for model, count in report["models_in_xlsx"].most_common():
        norm = normalize_str(model)
        matched = product_by_code.get(norm) or product_by_name.get(norm)
        bucket = "models_matched" if matched else "models_unmatched"
        report[bucket].append({
            "model": model,
            "deal_count": count,
            "product_id": matched["id"] if matched else None,
            "product_code": matched.get("model_code") if matched else None,
        })

    # Coerce Counters to dicts for JSON
    report["deals_by_status"] = dict(report["deals_by_status"])
    report["reps_in_xlsx"] = dict(report["reps_in_xlsx"])
    report["models_in_xlsx"] = dict(report["models_in_xlsx"])

    # Write JSON
    out_json = "/tmp/backfill-dry-run-report.json"
    with open(out_json, "w") as f:
        json.dump(report, f, indent=2, default=str)

    # Write Markdown summary
    out_md = "/tmp/backfill-dry-run-report.md"
    with open(out_md, "w") as f:
        f.write(f"# Backfill Dry-Run Report\n\n")
        f.write(f"Generated: {report['generated_at']}\n\n")
        f.write(f"## Files\n\n")
        f.write(f"- Total XLSX files: **{report['xlsx_files_total']}**\n")
        f.write(f"- Unparseable: {len(report['unparseable_files'])}\n\n")
        f.write(f"## Shows\n\n")
        f.write(f"- Total shows in XLSX: **{len(report['shows_in_xlsx'])}**\n")
        f.write(f"- Matched cleanly to a Supabase show: **{report['shows_matched']}**\n")
        f.write(f"- No match (would create new): **{report['shows_unmatched']}**\n")
        f.write(f"- Ambiguous (multiple Supabase matches on same date+city): {report['shows_ambiguous']}\n\n")
        f.write(f"## Deals\n\n")
        f.write(f"- Total deal rows: **{report['deals_total']}**\n")
        for status, count in sorted(report["deals_by_status"].items(), key=lambda x: -x[1]):
            f.write(f"  - {status}: {count}\n")
        f.write(f"\n")
        f.write(f"## Sales Reps\n\n")
        f.write(f"- Unique reps in XLSX: **{len(report['reps_in_xlsx'])}**\n")
        f.write(f"- Matched to existing profiles: **{len(report['reps_matched'])}**\n")
        f.write(f"- Unmatched (would need new profiles): **{len(report['reps_unmatched'])}**\n\n")
        if report["reps_unmatched"]:
            f.write(f"### Top unmatched reps (by deal count)\n\n")
            for entry in report["reps_unmatched"][:15]:
                f.write(f"- {entry['name']} ({entry['deal_count']} deals)\n")
            f.write("\n")
        f.write(f"## Models\n\n")
        f.write(f"- Unique models in XLSX: **{len(report['models_in_xlsx'])}**\n")
        f.write(f"- Matched to existing products: **{len(report['models_matched'])}**\n")
        f.write(f"- Unmatched (would need lookup or 'Other'): **{len(report['models_unmatched'])}**\n\n")
        if report["models_unmatched"]:
            f.write(f"### Top unmatched models (by deal count)\n\n")
            for entry in report["models_unmatched"][:15]:
                f.write(f"- {entry['model']} ({entry['deal_count']} deals)\n")
            f.write("\n")
        if report["unparseable_files"]:
            f.write(f"## Unparseable files\n\n")
            for fn in report["unparseable_files"]:
                f.write(f"- {fn}\n")

    print(f"\nReport written:\n  {out_json}\n  {out_md}", file=sys.stderr)


if __name__ == "__main__":
    main()
