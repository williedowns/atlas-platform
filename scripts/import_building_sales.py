#!/usr/bin/env python3
"""
Import Atlas Building Systems sales from Lori's XLSXs into Supabase.

Parses three different workbook structures:
  1. 2026 WEEKLY SALES.xlsx  — MASTER + weekly + monthly-total tabs (cleanest schema)
  2. 2025 WEEKLY SALES.xlsx  — Running Total tab (transactional) + WH-SALES (wholesale)
  3. 2025 SALES.xlsx         — All Stores tab (full year transactional)

Reconciles to one truth: each transaction is identified by a row_hash
(MD5 of file/sheet/row/date/amount/product). Duplicate hashes are skipped on
re-import via the unique index.

Usage:
    # Inspect what would be inserted, no DB writes:
    python3 scripts/import_building_sales.py --dry-run

    # Insert into Supabase:
    python3 scripts/import_building_sales.py --execute

    # Parse just one file (for debugging):
    python3 scripts/import_building_sales.py --dry-run --only "2026 WEEKLY SALES.xlsx"

Environment:
    NEXT_PUBLIC_SUPABASE_URL          (from .env.local)
    SUPABASE_SERVICE_ROLE_KEY         (from .env.local)
"""

import argparse
import hashlib
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Default file locations ────────────────────────────────────────────────
DEFAULT_FILES = [
    "/Users/williedowns/Downloads/2026 WEEKLY SALES.xlsx",
    "/Users/williedowns/Downloads/2025 WEEKLY SALES.xlsx",
    "/Users/williedowns/Downloads/2025 SALES.xlsx",
]

# Location names get normalized so "Plano Spa", "PLANO", "Plano" all fold
# into a canonical key. The canonical name then matches what we'll FK
# against the locations table later.
LOCATION_NORM = {
    "canton": "Canton",
    "ennis": "Ennis",
    "fort worth": "Fort Worth",
    "fort woth": "Fort Worth",       # XLSX typo
    "fort worth atlas": "Fort Worth",
    "fort woth spa": "Fort Worth",
    "glenn heights": "Glenn Heights",
    "longview": "Longview",
    "lufkin": "Lufkin",
    "palestine": "Palestine",
    "palestine i": "Palestine",
    "palestine ii": "Palestine II",
    "plano": "Plano",
    "plano spa": "Plano",
    "plano spa / atlas": "Plano",
    "plano spa  atlas": "Plano",
    "tyler": "Tyler",
    "tyler spa": "Tyler",
    "tyler spa / atlas": "Tyler",
    "tyler spa  atlas": "Tyler",
    "waco": "Waco",
    "waco spa": "Waco",
    "waco spa / atlas": "Waco",
    "waco spa  atlas": "Waco",
    "service": "Service",
    "company": "Company",
    "okc": "Oklahoma City",
    "oklahoma city": "Oklahoma City",
    "oklahoma city / pools": "Oklahoma City",
    "okc / pools": "Oklahoma City",
    "wh / athens": "WH — Athens",
    "wh athens": "WH — Athens",
    "wh / canton": "WH — Canton",
    "wh canton": "WH — Canton",
}

# Buildings/Atlas product taxonomy. Anything matching is a building sale;
# spa-specific products (chemicals, 110V Cord, 50 Amp Breaker, board parts)
# are excluded — they belong to the Atlas Spas division.
BUILDING_KEYWORDS = re.compile(
    r"\b(barn|deluxe|economy|classic|supreme|bargain|carport|cabin|deck|"
    r"granite\s*base|concrete|pool|sentinelle|garage|shed|cottage|"
    r"playhouse|porch|side\s*lofted|lofted|chicken\s*coop)\b",
    re.IGNORECASE,
)
SPA_ONLY_KEYWORDS = re.compile(
    r"\b(chemicals?|110v|220v|amp|breaker|board|filter|cover|ozonator|"
    r"jet|pillow|spa\s*pack|cord|fuse|relay|thermistor|topside|grommet|"
    r"o[\-\s]?ring|gasket|valve|union|pump|heater)\b",
    re.IGNORECASE,
)


def normalize_location(raw: str) -> Optional[str]:
    """Fold variants ('Plano Spa', 'PLANO', 'Plano') to canonical form.
    Any string starting with 'WH' (wholesale) is preserved with that prefix
    so the channel classifier below treats it correctly."""
    if not raw:
        return None
    s = str(raw).strip()
    if not s:
        return None
    key = re.sub(r"\s+", " ", s.lower())
    # Direct map hit
    if key in LOCATION_NORM:
        return LOCATION_NORM[key]
    # Wholesale fall-through — preserve "WH — <suffix>" canonical form
    if key.startswith("wh ") or key.startswith("wh/") or " wh " in f" {key} ":
        # Strip leading "WH /" or "WH" and clean up
        rest = re.sub(r"^wh\s*[/\-]?\s*", "", key).strip()
        # Title-case but keep Roman numerals upper (II, III, IV) and short
        # acronyms (Harris stays Harris, not "Iii")
        def cap(part: str) -> str:
            if re.fullmatch(r"i{1,4}|v|x", part):
                return part.upper()
            return part.capitalize()
        rest_title = " ".join(cap(p) for p in re.split(r"\s+|/", rest) if p)
        return f"WH — {rest_title}" if rest_title else "WH — Unknown"
    return s


def is_wholesale_location(loc_name: str) -> bool:
    """Wholesale = any normalized location starting with our WH marker."""
    return bool(loc_name) and loc_name.startswith("WH ")


def is_building_product(product: str) -> bool:
    """True if product matches the buildings taxonomy."""
    if not product:
        return False
    p = str(product).strip()
    if not p:
        return False
    # Explicit spa-only exclusions
    if SPA_ONLY_KEYWORDS.search(p) and not BUILDING_KEYWORDS.search(p):
        return False
    if BUILDING_KEYWORDS.search(p):
        return True
    return False


def to_date(v) -> Optional[date]:
    if v is None:
        return None
    d = None
    if isinstance(v, datetime):
        d = v.date()
    elif isinstance(v, date):
        d = v
    else:
        try:
            d = datetime.fromisoformat(str(v)).date()
        except Exception:
            return None
    # Filter implausible dates — XLSX sometimes has '2002' typos for '2025'
    if d.year < 2024 or d.year > 2030:
        return None
    return d


def to_amount(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace("$", "").replace(",", "").strip()
        return float(s) if s else None
    except Exception:
        return None


def row_hash(file_basename: str, sheet: str, row_idx: int, sold_at, amount, product) -> str:
    h = hashlib.md5()
    h.update(f"{file_basename}|{sheet}|{row_idx}|{sold_at}|{amount}|{product}".encode())
    return h.hexdigest()


# ── Parsers ───────────────────────────────────────────────────────────────

def parse_2025_sales(path: str):
    """Parse '2025 SALES.xlsx' — annual transaction log.
    Use the 'All Stores' tab (4055 rows) which is the master.
    Columns: A=Date, B=Store, C=Product, D=Amount, E=Size"""
    file_basename = Path(path).name
    wb = openpyxl.load_workbook(path, data_only=True)
    if "All Stores" not in wb.sheetnames:
        return
    ws = wb["All Stores"]
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None for c in row):
            continue
        sold_at = to_date(row[0])
        store = row[1] if len(row) > 1 else None
        product = row[2] if len(row) > 2 else None
        amount = to_amount(row[3] if len(row) > 3 else None)
        size = row[4] if len(row) > 4 else None

        if not sold_at or amount is None or not product:
            continue
        if not is_building_product(product):
            continue

        loc = normalize_location(store)
        if not loc:
            continue

        yield {
            "sold_at": sold_at,
            "location_name": loc,
            "product_category": str(product).strip(),
            "product_size": str(size).strip() if size else None,
            "amount": amount,
            "stock_status": None,
            "channel": "wholesale" if is_wholesale_location(loc) else "retail",
            "salesman_name": None,
            "source_file": file_basename,
            "source_row_hash": row_hash(file_basename, "All Stores", r_idx, sold_at, amount, product),
        }
    wb.close()


def parse_2025_weekly(path: str):
    """Parse '2025 WEEKLY SALES.xlsx' — use the 'Running Total' tab (4015 rows)
    which is transactional and cleaner than the weekly grids.
    Plus 'WH - SALES ' tab for wholesale rows.
    Schema: A=Date, B=Store, C=Product, D=Amount, E=Size"""
    file_basename = Path(path).name
    wb = openpyxl.load_workbook(path, data_only=True)

    for sheet_name, channel in [("Running Total", "retail"), ("WH - SALES ", "wholesale")]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            sold_at = to_date(row[0] if len(row) > 0 else None)
            store = row[1] if len(row) > 1 else None
            product = row[2] if len(row) > 2 else None
            amount = to_amount(row[3] if len(row) > 3 else None)
            size = row[4] if len(row) > 4 else None
            if not sold_at or amount is None or not product:
                continue
            if not is_building_product(product):
                continue
            loc = normalize_location(store)
            if not loc:
                continue
            yield {
                "sold_at": sold_at,
                "location_name": loc,
                "product_category": str(product).strip(),
                "product_size": str(size).strip() if size else None,
                "amount": amount,
                "stock_status": None,
                "channel": channel,
                "salesman_name": None,
                "source_file": file_basename,
                "source_row_hash": row_hash(file_basename, sheet_name, r_idx, sold_at, amount, product),
            }
    wb.close()


def parse_2026_weekly(path: str):
    """Parse '2026 WEEKLY SALES.xlsx' — schema: A=Date, B=Product, C=Details,
    D=Sold For, E=Stock status, F=Location. Note DIFFERENT column order from
    2025 files. The 'MASTER' tab is a template (empty data rows) so we
    iterate ALL weekly sheets — sheet names like '5-18 to 5-23'."""
    file_basename = Path(path).name
    wb = openpyxl.load_workbook(path, data_only=True)
    # Skip pivots/summaries; only sheets that look like weekly data:
    SKIP = {"MASTER", "Monthly", "2025 SALES", "Test for formulas",
            "January Total", "February Total", "March Total", "April Totals"}
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP:
            continue
        ws = wb[sheet_name]
        # Verify schema by checking row 4 — should have DATE / PRODUCT / etc.
        # Some weekly sheets start at different rows; use a flexible detector.
        header_row = None
        for try_row in range(1, 8):
            cells = [str(c.value or "").strip().upper() for c in ws[try_row]]
            if cells and "DATE" in cells[:2] and "PRODUCT" in cells[:3]:
                header_row = try_row
                break
        if header_row is None:
            continue  # not a weekly schema sheet
        for r_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not row or all(c is None for c in row):
                continue
            sold_at = to_date(row[0] if len(row) > 0 else None)
            product = row[1] if len(row) > 1 else None
            size = row[2] if len(row) > 2 else None
            amount = to_amount(row[3] if len(row) > 3 else None)
            stock_status = row[4] if len(row) > 4 else None
            store = row[5] if len(row) > 5 else None
            if not sold_at or amount is None or not product:
                continue
            if not is_building_product(product):
                continue
            loc = normalize_location(store)
            if not loc:
                continue
            yield {
                "sold_at": sold_at,
                "location_name": loc,
                "product_category": str(product).strip(),
                "product_size": str(size).strip() if size else None,
                "amount": amount,
                "stock_status": str(stock_status).strip() if stock_status else None,
                "channel": "wholesale" if is_wholesale_location(loc) else "retail",
                "salesman_name": None,
                "source_file": file_basename,
                "source_row_hash": row_hash(file_basename, sheet_name, r_idx, sold_at, amount, product),
            }
    wb.close()


PARSERS = [
    # 2025 WEEKLY's "Running Total" tab covers all of 2025 transactionally
    # and is actively maintained. 2025 SALES.xlsx "All Stores" tab duplicates
    # the same transactions (Canton showed 2x in dry-run). Skip the SALES
    # file by default to avoid double-counting; if you need it later, the
    # cross-file dedup question gets revisited.
    ("2026 WEEKLY SALES.xlsx", parse_2026_weekly),
    ("2025 WEEKLY SALES.xlsx", parse_2025_weekly),
]


def parse_all(files, only=None):
    """Yield canonical rows from all configured parsers."""
    by_name = {Path(f).name: f for f in files}
    for filename, parser in PARSERS:
        if only and only != filename:
            continue
        if filename not in by_name:
            print(f"  [skip] {filename} — file not found", file=sys.stderr)
            continue
        path = by_name[filename]
        print(f"\n  parsing {filename}…")
        yield from parser(path)


# ── Main ──────────────────────────────────────────────────────────────────

def summarize(rows):
    """Print a summary of parsed rows for dry-run verification."""
    by_year_loc = defaultdict(lambda: defaultdict(float))
    by_year_total = defaultdict(float)
    by_channel = defaultdict(float)
    by_file = defaultdict(int)
    for r in rows:
        y = r["sold_at"].year
        by_year_loc[y][r["location_name"]] += r["amount"]
        by_year_total[y] += r["amount"]
        by_channel[r["channel"]] += r["amount"]
        by_file[r["source_file"]] += 1
    print("\n──────────────────────────────────────────────────")
    print(f"Total rows parsed: {len(rows)}")
    print("\nBy source file:")
    for f, n in sorted(by_file.items()):
        print(f"  {f:40s} {n:6d} rows")
    print("\nBy year × location (top 10 per year):")
    for y in sorted(by_year_loc):
        print(f"\n  {y}:  ${by_year_total[y]:>12,.2f} total")
        top = sorted(by_year_loc[y].items(), key=lambda kv: -kv[1])[:10]
        for loc, amt in top:
            print(f"    {loc:30s} ${amt:>12,.2f}")
    print("\nBy channel:")
    for ch, amt in sorted(by_channel.items()):
        print(f"  {ch:12s} ${amt:>14,.2f}")
    print("──────────────────────────────────────────────────")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="Parse + summarize, no DB writes")
    ap.add_argument("--execute", action="store_true", help="Actually insert into Supabase")
    ap.add_argument("--only", type=str, default=None, help="Process only this XLSX (filename)")
    ap.add_argument("--files", nargs="*", default=DEFAULT_FILES, help="Override default file paths")
    args = ap.parse_args()

    if not args.dry_run and not args.execute:
        print("ERROR: pass --dry-run or --execute", file=sys.stderr)
        sys.exit(2)

    rows = list(parse_all(args.files, only=args.only))

    if args.dry_run:
        summarize(rows)
        print("\nDry run complete. Re-run with --execute to insert.")
        return

    # --execute path
    try:
        from supabase import create_client
    except ImportError:
        print("ERROR: supabase-py not installed. Run: pip3 install supabase", file=sys.stderr)
        sys.exit(1)

    # Load .env.local
    env_path = Path(__file__).parent.parent / ".env.local"
    env = {}
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            m = re.match(r"^([A-Z_]+)\s*=\s*(.*?)\s*$", line)
            if m:
                env[m.group(1)] = m.group(2).strip('"\'')

    url = env.get("NEXT_PUBLIC_SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY", file=sys.stderr)
        sys.exit(1)

    sb = create_client(url, key)

    # Insert in chunks. The unique index on (source_file, source_row_hash)
    # makes duplicates a no-op via ON CONFLICT DO NOTHING.
    CHUNK = 200
    inserted = 0
    skipped = 0
    for i in range(0, len(rows), CHUNK):
        chunk = rows[i:i + CHUNK]
        payload = [
            {**r, "sold_at": r["sold_at"].isoformat()}
            for r in chunk
        ]
        result = sb.table("building_sales").upsert(
            payload,
            on_conflict="source_file,source_row_hash",
            ignore_duplicates=True,
        ).execute()
        inserted += len(result.data or [])

    print(f"\n  Insert complete: {inserted} new rows. (Duplicates auto-skipped.)")


if __name__ == "__main__":
    main()
