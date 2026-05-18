#!/usr/bin/env python3
"""
Create the show-sales XLSX template by taking one of Lori Donahue's actual
expo workbooks and stripping out the customer data rows while preserving
ALL sheets, formulas, formatting, charts, named ranges, and cross-sheet
references.

The TypeScript export module (src/lib/show-sales/xlsx-export.ts) loads
this template, injects deal data into the Sales tab starting at row 4,
populates show config into the Variables tab, and saves the result.

Usage:
    python3 scripts/build_show_sales_template.py <source.xlsx>

Outputs:
    src/lib/show-sales/show-sales-template.xlsx
    src/lib/show-sales/template-meta.json   (column map, sheet inventory)
"""

import json
import os
import sys
from copy import copy

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


def strip_sales_data(ws, first_data_row=4):
    """Clear customer-identifying data from Sales tab rows >= first_data_row
    while preserving cell formats, formulas, and styles by re-applying them
    after clearing the value."""
    max_row = ws.max_row
    max_col = ws.max_column
    cleared = 0
    for row_idx in range(first_data_row, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Preserve formulas (cells whose value starts with '=' AND that
            # appear consistently across rows are template formulas, not user
            # input). Heuristic: if it's a formula referencing the same row,
            # keep it. Otherwise clear.
            if cell.value is None:
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                # Keep formulas — they are the template's business logic
                continue
            # Clear the data value but leave format/style intact
            cell.value = None
            cleared += 1
    return cleared


def build_meta(wb):
    """Produce a JSON document that describes the template's structure so
    the TypeScript export module knows where to inject data."""
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "frozen_panes": ws.freeze_panes,
        }
        if sheet_name == "Sales":
            # Capture the column header → column letter map for Sales
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                letter = get_column_letter(col_idx)
                headers.append({
                    "col": col_idx,
                    "letter": letter,
                    "header": header,
                })
            sheet_info["headers"] = headers
            sheet_info["first_data_row"] = 4
        sheets.append(sheet_info)
    return {"sheets": sheets}


def main():
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        sys.exit(1)
    src = sys.argv[1]
    if not os.path.isfile(src):
        print(f"ERROR: source file not found: {src}", file=sys.stderr)
        sys.exit(2)

    here = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.normpath(os.path.join(here, "..", "src", "lib", "show-sales"))
    os.makedirs(out_dir, exist_ok=True)

    out_xlsx = os.path.join(out_dir, "show-sales-template.xlsx")
    out_meta = os.path.join(out_dir, "template-meta.json")

    print(f"Loading {src}", file=sys.stderr)
    # keep_vba=False since these aren't macro workbooks; data_only=False
    # preserves formulas in the saved file
    wb = openpyxl.load_workbook(src, data_only=False)

    print(f"Sheets: {wb.sheetnames}", file=sys.stderr)

    if "Sales" not in wb.sheetnames:
        print(f"ERROR: source workbook has no 'Sales' sheet", file=sys.stderr)
        sys.exit(3)

    cleared = strip_sales_data(wb["Sales"])
    print(f"Cleared {cleared} customer-data cells from Sales tab (preserved formulas)", file=sys.stderr)

    meta = build_meta(wb)

    wb.save(out_xlsx)
    print(f"Wrote template → {out_xlsx}", file=sys.stderr)

    with open(out_meta, "w") as fh:
        json.dump(meta, fh, indent=2)
    print(f"Wrote meta → {out_meta}", file=sys.stderr)


if __name__ == "__main__":
    main()
