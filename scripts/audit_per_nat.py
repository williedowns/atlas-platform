#!/usr/bin/env python3
"""
Per Nat audit:
  1. Reads Per NatTim sheet of "Per Nat.xlsx" — 454 deals.
  2. Pulls every contract from production with customer name.
  3. Fuzzy-matches XLSX rows → contracts.
  4. Reports: matched & flagged, matched & NOT flagged (need backfill),
     unmatched (no DB record).
  5. Emits a SQL file with UPDATE statements for the matched-but-not-flagged
     set, copying the XLSX timeframe + notes + fierce_notes into the
     contract.
"""

import os
import re
import sys
import json
import warnings
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from supabase import create_client
except ImportError as e:
    print(f"ERROR: {e}", file=sys.stderr)
    sys.exit(1)

warnings.filterwarnings("ignore")

XLSX_PATH = "/Users/williedowns/Downloads/Per Nat.xlsx"
OUTPUT_DIR = Path("/tmp/per_nat_audit")
OUTPUT_DIR.mkdir(exist_ok=True)


def norm(s):
    """Aggressive normalization for fuzzy match: lowercase, strip non-alnum."""
    if not s:
        return ""
    return re.sub(r"[^a-z0-9]", "", s.lower())


def name_keys(raw):
    """
    XLSX customer names are freeform: 'Charles & Stephen Gustav Bostelman',
    'Tim/Sheryl Frank', 'Kathleen & Scott Daum'. Return a set of fuzzy
    keys we'd accept as matches — typically last-name tokens.
    """
    if not raw:
        return set()
    # Strip parens, normalize separators to space
    s = re.sub(r"\([^)]*\)", " ", raw)
    s = re.sub(r"[/&,]", " ", s)
    # Tokenize on whitespace
    tokens = [t for t in s.split() if len(t) > 1]
    if not tokens:
        return set()
    # The last token is usually the last name. Sometimes two-word last names
    # like "Van Dyne" — include the last 2 tokens combined as a fallback.
    keys = set()
    if len(tokens) >= 1:
        keys.add(norm(tokens[-1]))
    if len(tokens) >= 2:
        keys.add(norm(tokens[-2] + tokens[-1]))
    return keys


def load_per_nat_rows():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Per NatTim"]
    rows = []
    for r in range(2, ws.max_row + 1):
        date_v = ws.cell(r, 1).value
        timeframe = ws.cell(r, 2).value
        customer = ws.cell(r, 3).value
        serial = ws.cell(r, 4).value
        model = ws.cell(r, 5).value
        color = ws.cell(r, 6).value
        skirt = ws.cell(r, 7).value
        notes = ws.cell(r, 8).value
        fierce = ws.cell(r, 10).value
        if not customer or not isinstance(customer, str):
            continue
        cn = customer.strip()
        if len(cn) < 3 or cn.startswith("="):
            continue
        rows.append({
            "date": str(date_v) if date_v else None,
            "timeframe": str(timeframe).strip() if timeframe else None,
            "customer": cn,
            "name_keys": list(name_keys(cn)),
            "serial": str(serial).strip() if serial else None,
            "model": str(model).strip() if model else None,
            "color": str(color).strip() if color else None,
            "skirt": str(skirt).strip() if skirt else None,
            "notes": str(notes).strip() if notes else None,
            "fierce_notes": str(fierce).strip() if fierce else None,
        })
    return rows


def fetch_contracts(sb):
    """Pull every contract with customer name and Per Nat flag state."""
    contracts = []
    page_size = 1000
    start = 0
    while True:
        result = (
            sb.table("contracts")
            .select(
                "id, contract_number, status, is_per_nat, per_nat_reason, "
                "delivery_timeframe, total, notes, created_at, "
                "customer:customers(first_name, last_name)"
            )
            .order("created_at", desc=True)
            .range(start, start + page_size - 1)
            .execute()
        )
        batch = result.data or []
        contracts.extend(batch)
        if len(batch) < page_size:
            break
        start += page_size
    # Pre-compute name keys
    for c in contracts:
        cust = c.get("customer") or {}
        first = (cust.get("first_name") or "").strip()
        last = (cust.get("last_name") or "").strip()
        c["_first"] = first
        c["_last"] = last
        c["_keys"] = set()
        if last:
            c["_keys"].add(norm(last))
        if first and last:
            c["_keys"].add(norm(first + last))
    return contracts


def main():
    # Load env from .env.local
    env_path = Path(__file__).parent.parent / ".env.local"
    for line in env_path.read_text().splitlines():
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

    sb = create_client(
        os.environ["NEXT_PUBLIC_SUPABASE_URL"],
        os.environ["SUPABASE_SERVICE_ROLE_KEY"],
    )

    print("=== Per Nat audit ===")
    print(f"Loading XLSX from {XLSX_PATH}...")
    xlsx_rows = load_per_nat_rows()
    print(f"  {len(xlsx_rows)} Per Nat deals in XLSX")

    print("Fetching contracts from production...")
    contracts = fetch_contracts(sb)
    print(f"  {len(contracts)} contracts in DB")

    flagged_in_db = sum(1 for c in contracts if c.get("is_per_nat"))
    print(f"  {flagged_in_db} currently flagged is_per_nat=true")

    # Match XLSX rows to contracts
    matched_flagged = []
    matched_unflagged = []
    unmatched = []

    by_key = {}
    for c in contracts:
        for k in c["_keys"]:
            by_key.setdefault(k, []).append(c)

    for row in xlsx_rows:
        candidates = []
        for k in row["name_keys"]:
            candidates.extend(by_key.get(k, []))
        # Deduplicate
        seen = set()
        uniq = []
        for c in candidates:
            if c["id"] not in seen:
                seen.add(c["id"])
                uniq.append(c)
        if not uniq:
            unmatched.append(row)
            continue
        # Prefer non-cancelled, non-delivered candidates that aren't already flagged
        uniq.sort(key=lambda c: (
            c.get("status") in ("cancelled", "delivered"),
            bool(c.get("is_per_nat")),
            -(c.get("total") or 0),
        ))
        match = uniq[0]
        row["_match_id"] = match["id"]
        row["_match_contract_number"] = match["contract_number"]
        row["_match_status"] = match["status"]
        if match.get("is_per_nat"):
            matched_flagged.append(row)
        else:
            matched_unflagged.append(row)

    print()
    print(f"Matched & already flagged:     {len(matched_flagged):3d}")
    print(f"Matched & NOT flagged (BACKFILL NEEDED): {len(matched_unflagged):3d}")
    print(f"Unmatched (no DB record):      {len(unmatched):3d}")
    print()

    # Write SQL backfill for the unflagged matches
    sql_path = OUTPUT_DIR / "per_nat_backfill.sql"
    with open(sql_path, "w") as f:
        f.write("-- Per Nat backfill — flag matched contracts and copy XLSX detail.\n")
        f.write("-- Generated by scripts/audit_per_nat.py\n")
        f.write("-- Source: /Users/williedowns/Downloads/Per Nat.xlsx (Per NatTim sheet)\n")
        f.write(f"-- Generated at: {datetime.utcnow().isoformat()}Z\n")
        f.write("-- Rows below: matched-but-not-flagged contracts.\n")
        f.write("--\n")
        f.write("-- For each row: set is_per_nat=true, per_nat_reason='manual',\n")
        f.write("-- and copy XLSX timeframe + notes + fierce_notes onto the contract.\n")
        f.write("-- We use per_nat_reason='manual' because these don't all have\n")
        f.write("-- status='low_deposit' or delivery_timeframe set today; Natalie\n")
        f.write("-- put them on the Per Nat list for various reasons captured in\n")
        f.write("-- the notes column.\n")
        f.write("\n")
        f.write("BEGIN;\n\n")
        for row in matched_unflagged:
            cid = row["_match_id"]
            tf = row.get("timeframe") or ""
            notes = (row.get("notes") or "").replace("'", "''")
            fierce = (row.get("fierce_notes") or "").replace("'", "''")
            tf_sql = "NULL" if not tf else f"'{tf.replace(chr(39), chr(39)+chr(39))}'"
            # Compose contract.notes by appending Per Nat detail.
            new_notes_parts = []
            if notes:
                new_notes_parts.append(f"Per Nat: {notes}")
            if fierce:
                new_notes_parts.append(f"Delivery: {fierce}")
            new_notes = "\n\n".join(new_notes_parts)
            new_notes_sql = "NULL" if not new_notes else f"'{new_notes}'"
            f.write(f"UPDATE public.contracts SET\n")
            f.write(f"  is_per_nat = true,\n")
            f.write(f"  per_nat_reason = COALESCE(per_nat_reason, 'manual'),\n")
            f.write(f"  delivery_timeframe = COALESCE(delivery_timeframe, {tf_sql}),\n")
            f.write(f"  notes = CASE\n")
            f.write(f"    WHEN notes IS NULL OR notes LIKE '[backfill%' THEN {new_notes_sql}\n")
            f.write(f"    ELSE notes\n")
            f.write(f"  END\n")
            f.write(f"WHERE id = '{cid}';\n")
            f.write(f"  -- {row['customer']} — {row.get('model') or ''}\n\n")
        f.write("\nCOMMIT;\n")
    print(f"Backfill SQL written: {sql_path}")
    print(f"  {len(matched_unflagged)} UPDATE statements")

    # Dump unmatched list
    unmatched_path = OUTPUT_DIR / "per_nat_unmatched.json"
    with open(unmatched_path, "w") as f:
        json.dump(unmatched, f, indent=2, default=str)
    print(f"Unmatched rows written: {unmatched_path}")

    # Summary report
    print()
    print("=== Summary ===")
    print(f"XLSX active Per Nat deals: {len(xlsx_rows)}")
    print(f"DB currently flags:        {flagged_in_db}")
    print(f"After backfill (estimate): {flagged_in_db + len(matched_unflagged)}")
    print(f"Still unmatched:           {len(unmatched)} (need data backfill from XLSX → contracts)")


if __name__ == "__main__":
    main()
