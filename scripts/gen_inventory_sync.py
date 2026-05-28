#!/usr/bin/env python3
"""Generate inventory_sync.sql from the Master XLSX.

Reads `~/Downloads/Hot Tub & Swim Spa Inventory.xlsx` (or path passed as $1),
emits `scripts/inventory_sync.sql` ready for the Atlas Spas POS Supabase SQL
Editor.

Upsert semantics:
  - Serialized units: ON CONFLICT (serial_number) DO UPDATE
  - On-order units (W-prefix order_number, no serial): ON CONFLICT (order_number) DO UPDATE
    backed by a partial unique index created at the top of the SQL.

Tab → status / location mapping is at the top of this file.
"""

from __future__ import annotations

import os
import sys
import warnings
from datetime import date, datetime
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

# ── Configuration ────────────────────────────────────────────────────────────

DEFAULT_XLSX = Path.home() / "Downloads" / "Hot Tub & Swim Spa Inventory.xlsx"
OUT_PATH = Path(__file__).parent / "inventory_sync.sql"
OUT_ACTIVE_PATH    = Path(__file__).parent / "inventory_sync_active.sql"
OUT_DELIVERED_PATH = Path(__file__).parent / "inventory_sync_delivered.sql"

META_TABS = {"KEY", "Status", "Settings", "Shell"}

# Tabs where col 0 holds a SHOW NAME header (and the same show name is repeated
# as col-0 of every row beneath it as a "marker"), rather than a customer name.
# In these tabs col 12 holds a HOME SHOWROOM name (e.g. "Tyler"), not a finance
# balance. Special-cased so we don't pollute customer_name / fin_balance.
SHOW_TABS = {"Expo 1", "Expo 2", "Expo 3", "Expo 4", "Expo 5",
             "Canton", "State Fair"}

# Known home-showroom labels that may appear in col-12 of Show tabs. If col-12
# matches one of these (case-insensitive), it's a home-showroom marker, not a
# balance.
HOME_SHOWROOM_LABELS = {
    "ennis", "tyler", "waco", "kansas", "okc", "georgetown",
    "plano", "houston", "ftw", "fort worth",
}

# Tab → (location_name, default_status)
# default_status is used when the XLSX Status column is blank or doesn't override.
TAB_MAP: dict[str, tuple[str, str]] = {
    # Physical showrooms
    "Ennis":      ("Ennis Warehouse",       "at_location"),
    "Tyler":      ("Tyler Showroom",       "at_location"),
    "Waco":       ("Waco Showroom",        "at_location"),
    "Kansas":     ("Kansas Showroom",      "at_location"),
    "OKC":        ("OKC Showroom",         "at_location"),
    "Georgetown": ("Georgetown Showroom",  "at_location"),
    "Plano":      ("Plano Showroom",       "at_location"),
    "Houston":    ("Houston Showroom",     "at_location"),
    "FTW":        ("Fort Worth Showroom",  "at_location"),
    # Special transit / build statuses
    "Take to Waco":   ("Waco Showroom",   "in_transit"),
    "Factory":        ("Ennis Warehouse",  "in_factory"),
    "Spas On Order":  ("Ennis Warehouse",  "on_order"),
    # Shows
    "Expo 1":      ("Ennis Warehouse", "at_show"),
    "Expo 2":      ("Ennis Warehouse", "at_show"),
    "Expo 3":      ("Ennis Warehouse", "at_show"),
    "Expo 4":      ("Ennis Warehouse", "at_show"),
    "Expo 5":      ("Ennis Warehouse", "at_show"),
    "Canton":      ("Ennis Warehouse", "at_show"),
    "State Fair":  ("Ennis Warehouse", "at_show"),
    # Historical
    "Delivered":   ("Ennis Warehouse", "delivered"),
}

# XLSX Status column value → DB status override
# When the XLSX row's Status column has one of these values, it overrides the
# tab default.
STATUS_OVERRIDES = {
    "sold":      "allocated",
    "delivered": "delivered",
    "pending":   None,     # use tab default
    "stock":     None,     # use tab default
}

# Column indices (0-based) — all tabs share this layout
COL_LAST_NAME    = 0
COL_FIRST_NAME   = 1
COL_WRAP         = 2
COL_LINE         = 3
COL_MODEL        = 4
COL_SHELL        = 5
COL_CABINET      = 6
COL_SERIAL       = 7   # or order # for on-order tabs
COL_LOCATION     = 8
COL_COMPLETED    = 9
COL_STATUS       = 10
COL_NOTES_1      = 11  # "Fierce Notes" / "Notes" / "Houston Notes" / etc.
COL_FIN_BAL      = 12
COL_ATLAS_NOTES  = 13
COL_EXPO_NOTES   = 14


# ── Helpers ──────────────────────────────────────────────────────────────────

def sql_escape(s: str | None) -> str:
    """SQL-quote a string (single quotes) or return NULL."""
    if s is None:
        return "NULL"
    s = str(s).strip()
    if not s:
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def is_blank(v) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s.lower() == "none"


def clean_serial(v) -> str | None:
    """Normalize serial: trim, drop trailing .0 for ints, uppercase letters."""
    if is_blank(v):
        return None
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def is_order_number(s: str) -> bool:
    """W-prefix order numbers vs. true serial numbers."""
    return s.upper().startswith("W") and any(c.isdigit() for c in s)


def normalize_wrap(v) -> str | None:
    if is_blank(v):
        return None
    s = str(v).strip().upper()
    if s in ("WR", "UN"):
        return s
    return None


def normalize_status(tab_default: str, raw_status: str | None) -> str:
    if is_blank(raw_status):
        return tab_default
    k = str(raw_status).strip().lower()
    if k in STATUS_OVERRIDES:
        ov = STATUS_OVERRIDES[k]
        return ov if ov else tab_default
    return tab_default


def build_customer_name(last, first) -> str | None:
    l = "" if is_blank(last) else str(last).strip()
    f = "" if is_blank(first) else str(first).strip()
    if not l and not f:
        return None
    if l and f:
        return f"{l}, {f}"
    return l or f


def parse_date(v) -> str | None:
    """Return ISO date string YYYY-MM-DD, or None if not a parseable date."""
    if is_blank(v):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None  # free-form text → ignore


def merge_notes(*chunks: tuple[str, object]) -> str | None:
    parts = []
    for label, val in chunks:
        if is_blank(val):
            continue
        if isinstance(val, datetime):
            text = val.date().isoformat()
        elif isinstance(val, date):
            text = val.isoformat()
        else:
            text = str(val).strip()
        parts.append(f"[{label}] {text}")
    if not parts:
        return None
    return " | ".join(parts)


# ── Main extraction ──────────────────────────────────────────────────────────

def extract_rows(xlsx_path: Path) -> tuple[list[dict], list[dict]]:
    """Returns (serialized_rows, on_order_rows)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    serialized: dict[str, dict] = {}   # serial_number → row
    on_order:   dict[str, dict] = {}   # order_number  → row

    # Process active tabs first, Delivered LAST. When a serial appears in both
    # an active tab and the historical Delivered tab, the active record wins —
    # the unit is still physically at that location, not yet delivered.
    sheet_order = [s for s in wb.sheetnames if s != "Delivered"]
    if "Delivered" in wb.sheetnames:
        sheet_order.append("Delivered")

    for sheet_name in sheet_order:
        if sheet_name in META_TABS:
            continue
        if sheet_name not in TAB_MAP:
            print(f"  [skip] {sheet_name}: not in TAB_MAP", file=sys.stderr)
            continue

        location_name, tab_default_status = TAB_MAP[sheet_name]
        ws = wb[sheet_name]
        is_show_tab = sheet_name in SHOW_TABS

        # For show tabs, the first row often has just the show name in col 0
        # (no serial) — capture it for use as a [Show] note on every unit below.
        tab_show_name: str | None = None
        if is_show_tab:
            for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                if row and row[0] and not clean_serial(row[COL_SERIAL]):
                    tab_show_name = str(row[0]).strip()
                    break

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 15:
                continue
            raw_key = clean_serial(row[COL_SERIAL])
            if not raw_key:
                continue

            status = normalize_status(tab_default_status, row[COL_STATUS])

            # Column 12 header varies ("Fin Bal" / "Est Comp") and worse, the value
            # type doesn't always match the header — Factory has "Fin Bal" header but
            # date values; OKC has "Est Comp" header but balance values. Route by
            # value type instead of header. (inventory_units has no approx_delivery_date
            # column — date values get folded into notes as [Est Comp] instead.)
            col12_val = row[COL_FIN_BAL]
            home_showroom_note: str | None = None
            est_comp_note: str | None = None
            if isinstance(col12_val, (datetime, date)):
                fin_balance = None
                est_comp_note = parse_date(col12_val)
            elif is_blank(col12_val):
                fin_balance = None
            else:
                fin_balance = str(col12_val).strip()
                # In Show tabs, col 12 is the unit's HOME SHOWROOM (e.g. "Tyler"),
                # not a balance. Route to a notes label instead.
                if is_show_tab and fin_balance.lower() in HOME_SHOWROOM_LABELS:
                    home_showroom_note = fin_balance
                    fin_balance = None

            # Customer name: in show tabs, col 0 is usually the show name. Only
            # treat it as a real customer when status is "Sold" AND the value
            # doesn't match the show header.
            customer_name = build_customer_name(row[COL_LAST_NAME], row[COL_FIRST_NAME])
            if is_show_tab and customer_name and tab_show_name:
                first_word = customer_name.split(",")[0].strip().lower()
                show_first = tab_show_name.split(",")[0].strip().lower()
                if first_word == show_first or first_word.startswith(show_first):
                    customer_name = None
                elif status not in ("allocated", "delivered"):
                    # Stock units at a show should never have a customer name —
                    # whatever's in col 0 is a sticky note, not a buyer.
                    customer_name = None

            # Build notes with show + home-showroom + est-comp metadata.
            note_chunks = [
                ("Fierce", row[COL_NOTES_1]),
                ("Atlas",  row[COL_ATLAS_NOTES]),
                ("Expo",   row[COL_EXPO_NOTES]),
            ]
            if is_show_tab and tab_show_name:
                note_chunks.insert(0, ("Show", tab_show_name))
            if home_showroom_note:
                note_chunks.insert(1 if is_show_tab else 0, ("Home", home_showroom_note))
            if est_comp_note:
                note_chunks.insert(0, ("Est Comp", est_comp_note))

            data = {
                "serial_number": None,
                "order_number":  None,
                "location_name": location_name,
                "status":        status,
                "model_code":    None if is_blank(row[COL_MODEL])   else str(row[COL_MODEL]).strip(),
                "shell_color":   None if is_blank(row[COL_SHELL])   else str(row[COL_SHELL]).strip(),
                "cabinet_color": None if is_blank(row[COL_CABINET]) else str(row[COL_CABINET]).strip(),
                "wrap_status":   normalize_wrap(row[COL_WRAP]),
                "customer_name": customer_name,
                "fin_balance":   fin_balance,
                "received_date": parse_date(row[COL_COMPLETED]),
                "notes":         merge_notes(*note_chunks),
                "_source_tab":   sheet_name,
            }

            if is_order_number(raw_key):
                # W-prefix => factory order number, never a serial regardless of status.
                # Force status to on_order unless the row is already Delivered (rare edge).
                if status not in ("delivered",):
                    data["status"] = "on_order"
                data["order_number"] = raw_key.upper()
                if raw_key.upper() not in on_order:
                    on_order[raw_key.upper()] = data   # first-wins (active before Delivered)
            else:
                data["serial_number"] = raw_key
                if raw_key not in serialized:
                    serialized[raw_key] = data         # first-wins (active before Delivered)

    wb.close()
    return list(serialized.values()), list(on_order.values())


# ── SQL emission ─────────────────────────────────────────────────────────────

def fmt_value_row(d: dict) -> str:
    """Produce one tuple for the VALUES list."""
    parts = [
        sql_escape(d["serial_number"]),
        sql_escape(d["order_number"]),
        sql_escape(d["location_name"]),
        sql_escape(d["status"]),
        sql_escape(d["model_code"]),
        sql_escape(d["shell_color"]),
        sql_escape(d["cabinet_color"]),
        sql_escape(d["wrap_status"]),
        sql_escape(d["customer_name"]),
        sql_escape(d["fin_balance"]),
        f"DATE {sql_escape(d['received_date'])}" if d["received_date"] else "NULL",
        sql_escape(d["notes"]),
    ]
    return "(" + ", ".join(parts) + ")"


COLUMNS = [
    "serial_number", "order_number", "location_id", "status",
    "model_code", "shell_color", "cabinet_color", "wrap_status",
    "customer_name", "fin_balance", "received_date", "notes",
]

SELECT_EXPRS = [
    "v.serial_number",
    "v.order_number",
    "(SELECT id FROM public.locations WHERE name = v.location_name LIMIT 1)",
    "v.status",
    "v.model_code",
    "v.shell_color",
    "v.cabinet_color",
    "v.wrap_status",
    "v.customer_name",
    "v.fin_balance",
    "v.received_date",
    "v.notes",
]

VALUES_COLUMNS = [
    "serial_number", "order_number", "location_name", "status",
    "model_code", "shell_color", "cabinet_color", "wrap_status",
    "customer_name", "fin_balance", "received_date", "notes",
]


def emit_sql(serialized: list[dict], on_order: list[dict], xlsx_path: Path) -> str:
    timestamp = datetime.now().isoformat(timespec="seconds")
    parts = [
        "-- ============================================================",
        f"-- Atlas Spas Inventory Sync — generated {timestamp}",
        f"-- Source: {xlsx_path}",
        f"-- Serialized units: {len(serialized)}",
        f"-- On-order units:   {len(on_order)}",
        "-- ============================================================",
        "",
        "BEGIN;",
        "",
        "-- Partial unique index for on-order units (no serial yet)",
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_inventory_units_order_number_unique",
        "  ON public.inventory_units(order_number)",
        "  WHERE serial_number IS NULL;",
        "",
    ]

    def emit_chunk(rows: list[dict], conflict_target: str, label: str) -> None:
        if not rows:
            return
        parts.append(f"-- ── {label} ({len(rows)} rows) ──")
        parts.append(
            f"INSERT INTO public.inventory_units ({', '.join(COLUMNS)})"
        )
        parts.append("SELECT")
        parts.append("  " + ",\n  ".join(SELECT_EXPRS))
        parts.append("FROM (")
        parts.append("  VALUES")
        value_rows = [fmt_value_row(r) for r in rows]
        parts.append(",\n".join("    " + v for v in value_rows))
        parts.append(f") AS v({', '.join(VALUES_COLUMNS)})")
        parts.append(f"ON CONFLICT ({conflict_target}) DO UPDATE SET")
        update_fields = [
            "location_id   = EXCLUDED.location_id",
            "status        = EXCLUDED.status",
            "model_code    = COALESCE(EXCLUDED.model_code, public.inventory_units.model_code)",
            "shell_color   = COALESCE(EXCLUDED.shell_color, public.inventory_units.shell_color)",
            "cabinet_color = COALESCE(EXCLUDED.cabinet_color, public.inventory_units.cabinet_color)",
            "wrap_status   = COALESCE(EXCLUDED.wrap_status, public.inventory_units.wrap_status)",
            "customer_name = EXCLUDED.customer_name",
            "fin_balance   = EXCLUDED.fin_balance",
            "received_date = COALESCE(EXCLUDED.received_date, public.inventory_units.received_date)",
            "notes         = EXCLUDED.notes",
            "updated_at    = now()",
        ]
        parts.append("  " + ",\n  ".join(update_fields) + ";")
        parts.append("")

    emit_chunk(serialized, "serial_number", "Serialized units")
    emit_chunk(on_order,   "order_number",  "On-order units (no serial)")

    parts.append("COMMIT;")
    parts.append("")
    parts.append(f"-- Totals: {len(serialized)} serialized + {len(on_order)} on-order = "
                 f"{len(serialized) + len(on_order)} rows")
    return "\n".join(parts)


# ── Entrypoint ───────────────────────────────────────────────────────────────

def main() -> None:
    xlsx_path = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"ERROR: XLSX not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading: {xlsx_path}", file=sys.stderr)
    serialized, on_order = extract_rows(xlsx_path)
    print(f"  serialized: {len(serialized)}", file=sys.stderr)
    print(f"  on-order:   {len(on_order)}", file=sys.stderr)

    # Per-tab breakdown for visibility
    from collections import Counter
    by_tab = Counter(r["_source_tab"] for r in serialized + on_order)
    for tab, n in sorted(by_tab.items(), key=lambda x: -x[1]):
        print(f"    {tab:18s} {n}", file=sys.stderr)

    # Combined file (everything)
    sql = emit_sql(serialized, on_order, xlsx_path)
    OUT_PATH.write_text(sql)
    print(f"\nWrote: {OUT_PATH}  ({OUT_PATH.stat().st_size:,} bytes)", file=sys.stderr)

    # Split files for Supabase SQL Editor (2.4 MB is too big to paste reliably).
    active   = [r for r in serialized if r["status"] != "delivered"]
    delivered = [r for r in serialized if r["status"] == "delivered"]

    sql_active = emit_sql(active, on_order, xlsx_path)
    OUT_ACTIVE_PATH.write_text(sql_active)
    print(f"Wrote: {OUT_ACTIVE_PATH}  ({OUT_ACTIVE_PATH.stat().st_size:,} bytes, "
          f"{len(active)} serialized + {len(on_order)} on-order)", file=sys.stderr)

    sql_delivered = emit_sql(delivered, [], xlsx_path)
    OUT_DELIVERED_PATH.write_text(sql_delivered)
    print(f"Wrote: {OUT_DELIVERED_PATH}  ({OUT_DELIVERED_PATH.stat().st_size:,} bytes, "
          f"{len(delivered)} delivered)", file=sys.stderr)


if __name__ == "__main__":
    main()
