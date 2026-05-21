#!/usr/bin/env python3
"""
Import all rows from Natalie's Per Nat XLSX into per_nat_entries.

Run AFTER migrations 085 + 086 are applied.

For each XLSX row:
  1. Track the current section as we walk the sheet (month dividers,
     "Owner to Notifiy YYYY" headers, black "Stock — held too long",
     red "HOT — can't reach customer").
  2. Fuzzy-match the customer name against contracts.customer.last_name.
  3. If matched → link contract_id.
  4. If unmatched → contract_id stays NULL (XLSX-only entry).
  5. Extract salesperson name from notes ("Tom's Deal.", "Conner Brady's
     Deal.") for the filter.

Idempotent — re-running clears per_nat_entries first.
"""

import os
import re
import sys
import warnings
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from supabase import create_client
except ImportError as e:
    print(f"ERROR: {e}", file=sys.stderr)
    sys.exit(1)

warnings.filterwarnings("ignore")

XLSX_PATH = "/Users/williedowns/Downloads/Per Nat.xlsx"
SALESPERSON_RX = re.compile(r"^([A-Z][a-zA-Z'\s\-]{1,30})\'s\s+Deal\b")

# XLSX section header fill colors (openpyxl reports as ARGB).
MONTH_DIVIDER_FILL = "FF0000FF"      # blue
OWNER_NOTIFY_FILL = "FF9900FF"       # purple/orange — "Owner to Notifiy YYYY"
STOCK_HELD_FILL = "FF000000"         # black — "Stock — held too long"
HOT_FILL = "FFFF0000"                # red — "HOT — can't reach customer"


def norm(s):
    if not s:
        return ""
    return re.sub(r"[^a-z0-9]", "", s.lower())


def name_keys(raw):
    if not raw:
        return set()
    s = re.sub(r"\([^)]*\)", " ", raw)
    s = re.sub(r"[/&,]", " ", s)
    tokens = [t for t in s.split() if len(t) > 1]
    keys = set()
    if len(tokens) >= 1:
        keys.add(norm(tokens[-1]))
    if len(tokens) >= 2:
        keys.add(norm(tokens[-2] + tokens[-1]))
    return keys


def extract_salesperson(notes):
    if not notes:
        return None
    m = SALESPERSON_RX.match(notes.strip())
    if m:
        return m.group(1).strip()
    return None


def parse_date(raw):
    if not raw:
        return None
    s = str(raw).strip()
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def cell_fill(cell):
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = cell.fill.fgColor.rgb
        if rgb in ("00000000", None):
            return None
        return rgb
    return None


def detect_section(ws, row_idx):
    """
    Returns (label, kind) if this row is a section header, else (None, None).

    A row is a section header when:
      - col A or col C has a recognized fill color,
      - some text is present, AND
      - the customer-name column (col C) is empty.

    The "Cancelled" sheet paints every data row red, so without the
    "no customer name" guard we'd treat every Cancelled deal as a section
    header.
    """
    customer = ws.cell(row_idx, 3).value
    if customer and isinstance(customer, str) and customer.strip():
        return None, None

    fill_a = cell_fill(ws.cell(row_idx, 1))
    fill_c = cell_fill(ws.cell(row_idx, 3))
    fill = fill_a or fill_c
    if not fill:
        return None, None

    # Gather header text (first non-empty cell, usually col A)
    text = None
    for c in range(1, 11):
        v = ws.cell(row_idx, c).value
        if v is None:
            continue
        if isinstance(v, datetime):
            text = v.strftime("%B %Y")
            break
        s = str(v).strip()
        if s:
            text = s
            break
    if not text:
        return None, None

    # Classify by TEXT first — the XLSX paints some "Owner to Notifiy"
    # rows with the same blue fill as month dividers, so fill alone is
    # ambiguous. Fall back to fill color when text doesn't match a known
    # categorical pattern.
    text_l = text.lower()
    kind = None
    if "owner to notif" in text_l:                              # "Owner to Notify YYYY" / "Owner to Notifiy YYYY"
        kind = "owner_notify"
    elif text_l.startswith("list below") or "mark as stock" in text_l:
        kind = "stock_held"
    elif text_l.startswith("hot") or text_l.startswith("don't forget") or text_l.startswith("dont forget"):
        kind = "hot"
    elif fill == MONTH_DIVIDER_FILL:
        kind = "month"
    elif fill == OWNER_NOTIFY_FILL:
        kind = "owner_notify"
    elif fill == STOCK_HELD_FILL:
        kind = "stock_held"
    elif fill == HOT_FILL:
        kind = "hot"
    else:
        return None, None

    # Clean month divider labels — "2026-05-01 00:00:00" -> "May 2026"
    if kind == "month":
        m = re.match(r"^(\d{4})-(\d{2})-\d{2}", text)
        if m:
            months = ["January","February","March","April","May","June","July","August","September","October","November","December"]
            year = m.group(1)
            mi = int(m.group(2)) - 1
            text = f"{months[mi]} {year}"

    return text, kind


def load_xlsx_rows(path, sheet_name, status):
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]

    rows = []
    current_section_label = None
    current_section_kind = None
    section_order_counter = 0

    for r in range(2, ws.max_row + 1):
        # First, check if this row is a section header.
        label, kind = detect_section(ws, r)
        if label is not None:
            section_order_counter += 1
            current_section_label = label
            current_section_kind = kind
            continue

        # Otherwise it's a data row — only include if customer name present.
        customer = ws.cell(r, 3).value
        if not customer or not isinstance(customer, str):
            continue
        cn = customer.strip()
        if len(cn) < 3 or cn.startswith("="):
            continue

        date_v = ws.cell(r, 1).value
        timeframe = ws.cell(r, 2).value
        serial = ws.cell(r, 4).value
        model = ws.cell(r, 5).value
        color = ws.cell(r, 6).value
        skirt = ws.cell(r, 7).value
        notes = ws.cell(r, 8).value
        fierce = ws.cell(r, 10).value

        rows.append({
            "sale_date": parse_date(date_v),
            "timeframe_text": str(timeframe).strip() if timeframe else None,
            "customer_name": cn,
            "_keys": list(name_keys(cn)),
            "serial_number": str(serial).strip() if serial else None,
            "model": str(model).strip() if model else None,
            "color": str(color).strip() if color else None,
            "skirt": str(skirt).strip() if skirt else None,
            "notes": str(notes).strip() if notes else None,
            "fierce_notes": str(fierce).strip() if fierce else None,
            "salesperson_name": extract_salesperson(str(notes) if notes else None),
            "status": status,
            "reason": "low_deposit" if notes and "low deposit" in notes.lower() else (
                "future_delivery" if timeframe else "manual"
            ),
            "source": "xlsx_import",
            "section_label": current_section_label,
            "section_kind": current_section_kind,
            "section_order": section_order_counter,
        })
    return rows


def fetch_contracts_index(sb):
    index = {}
    page_size = 1000
    start = 0
    while True:
        result = (
            sb.table("contracts")
            .select("id, status, total, customer:customers(first_name, last_name)")
            .range(start, start + page_size - 1)
            .execute()
        )
        batch = result.data or []
        for c in batch:
            cust = c.get("customer") or {}
            first = (cust.get("first_name") or "").strip()
            last = (cust.get("last_name") or "").strip()
            keys = set()
            if last:
                keys.add(norm(last))
            if first and last:
                keys.add(norm(first + last))
            for k in keys:
                index.setdefault(k, []).append(c)
        if len(batch) < page_size:
            break
        start += page_size
    return index


def main():
    env_path = Path(__file__).parent.parent / ".env.local"
    for line in env_path.read_text().splitlines():
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

    sb = create_client(
        os.environ["NEXT_PUBLIC_SUPABASE_URL"],
        os.environ["SUPABASE_SERVICE_ROLE_KEY"],
    )

    print("=== Per Nat XLSX import (with sections) ===")
    print("Loading XLSX sheets...")
    active = load_xlsx_rows(XLSX_PATH, "Per NatTim", "active")
    completed = load_xlsx_rows(XLSX_PATH, "Completed", "completed")
    cancelled = load_xlsx_rows(XLSX_PATH, "Cancelled", "cancelled")
    all_rows = active + completed + cancelled
    print(f"  Active: {len(active)}, Completed: {len(completed)}, Cancelled: {len(cancelled)}")
    print(f"  Total: {len(all_rows)}")

    # Show the detected section structure for the Active sheet so we can sanity-check.
    seen = set()
    print()
    print("Section structure detected in Per NatTim:")
    for r in active:
        key = (r["section_order"], r["section_label"], r["section_kind"])
        if key in seen:
            continue
        seen.add(key)
        if r["section_label"]:
            print(f"  [{r['section_order']:3d}] {r['section_kind']:12s} — {r['section_label']}")

    if "--dry-run" in sys.argv:
        print("\n--dry-run: not inserting.")
        return

    print()
    print("Indexing existing contracts...")
    index = fetch_contracts_index(sb)

    # Wipe ALL prior per_nat_entries — XLSX is the source of truth now.
    # contract_flag rows from migration 085's seed get replaced by xlsx_import
    # rows that match them (proper section + notes) OR by a second pass below
    # that re-adds contract_flag for is_per_nat=true contracts NOT in XLSX.
    print("Wiping per_nat_entries (XLSX is source of truth)...")
    sb.table("per_nat_entries").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()

    already_linked = set()

    matched = 0
    unmatched = 0
    inserts = []
    for row in all_rows:
        candidates = []
        for k in row["_keys"]:
            for c in index.get(k, []):
                if c["id"] in already_linked:
                    continue
                candidates.append(c)
        seen = set()
        uniq = []
        for c in candidates:
            if c["id"] not in seen:
                seen.add(c["id"])
                uniq.append(c)
        uniq.sort(key=lambda c: (
            c.get("status") in ("cancelled", "delivered"),
            -(c.get("total") or 0),
        ))

        link_id = None
        if uniq:
            link_id = uniq[0]["id"]
            already_linked.add(link_id)
            matched += 1
        else:
            unmatched += 1

        inserts.append({
            "contract_id": link_id,
            "source": row["source"],
            "sale_date": row["sale_date"],
            "customer_name": row["customer_name"],
            "model": row["model"],
            "color": row["color"],
            "skirt": row["skirt"],
            "serial_number": row["serial_number"],
            "salesperson_name": row["salesperson_name"],
            "timeframe_text": row["timeframe_text"],
            "notes": row["notes"],
            "fierce_notes": row["fierce_notes"],
            "status": row["status"],
            "reason": row["reason"],
            "section_label": row["section_label"],
            "section_kind": row["section_kind"],
            "section_order": row["section_order"],
        })

    print()
    print(f"  Matched to a contract: {matched}")
    print(f"  Unmatched (XLSX-only): {unmatched}")
    print()

    print(f"Inserting {len(inserts)} XLSX rows into per_nat_entries...")
    batch_size = 100
    inserted = 0
    for i in range(0, len(inserts), batch_size):
        chunk = inserts[i:i + batch_size]
        result = sb.table("per_nat_entries").insert(chunk).execute()
        inserted += len(result.data or [])
        print(f"  Batch {i // batch_size + 1}: {len(result.data or [])} rows (total: {inserted})")

    # ── Second pass: flagged-but-not-in-XLSX contracts ─────────────────────
    # Some contracts are is_per_nat=true (low_deposit auto-flag, or flagged
    # via the Modify Contract card) but never made it onto Natalie's XLSX.
    # Add them as contract_flag rows so the Per Nat page reflects the full
    # set, auto-assigned to a month section based on sale_date so they
    # don't pile up under TBD.
    print()
    print("Adding contract_flag entries for is_per_nat=true contracts not in XLSX...")
    flagged_result = (
        sb.table("contracts")
        .select(
            "id, status, total, deposit_paid, balance_due, "
            "per_nat_reason, delivery_timeframe, created_at, notes, "
            "customer:customers(first_name, last_name)"
        )
        .eq("is_per_nat", True)
        .execute()
    )
    flagged = flagged_result.data or []
    print(f"  {len(flagged)} contracts flagged is_per_nat=true")

    months = ["January","February","March","April","May","June","July","August","September","October","November","December"]
    pass2_inserts = []
    skipped_already_linked = 0
    for c in flagged:
        if c["id"] in already_linked:
            skipped_already_linked += 1
            continue

        # Status mapping
        cstatus = c.get("status") or ""
        if cstatus == "delivered":
            entry_status = "completed"
        elif cstatus == "cancelled":
            entry_status = "cancelled"
        else:
            entry_status = "active"

        # Auto-section by sale_date month so these rows don't land in TBD.
        sale_date_str = (c.get("created_at") or "")[:10]
        try:
            y, mo, _ = sale_date_str.split("-")
            section_label = f"{months[int(mo) - 1]} {y}"
            section_order = int(y) * 12 + int(mo) - 1
        except Exception:
            section_label = "TBD"
            section_order = 999999

        # Strip [backfill] noise from notes.
        raw_notes = c.get("notes") or ""
        clean_notes = re.sub(r"\[backfill[^\]]*\][^\n]*", "", raw_notes).strip() or None

        cust = c.get("customer") or {}
        customer_name = f"{(cust.get('first_name') or '').strip()} {(cust.get('last_name') or '').strip()}".strip() or "—"

        pass2_inserts.append({
            "contract_id": c["id"],
            "source": "contract_flag",
            "sale_date": sale_date_str or None,
            "customer_name": customer_name,
            "status": entry_status,
            "reason": c.get("per_nat_reason") or "manual",
            "timeframe_text": c.get("delivery_timeframe"),
            "notes": clean_notes,
            "section_label": section_label,
            "section_kind": "month",
            "section_order": section_order,
        })

    print(f"  Skipped (already linked from XLSX): {skipped_already_linked}")
    print(f"  Adding contract_flag rows: {len(pass2_inserts)}")
    if pass2_inserts:
        for i in range(0, len(pass2_inserts), batch_size):
            chunk = pass2_inserts[i:i + batch_size]
            sb.table("per_nat_entries").insert(chunk).execute()

    print()
    print(f"=== Done: {inserted + len(pass2_inserts)} entries total ===")
    print(f"    XLSX rows:        {inserted}")
    print(f"    contract_flag:    {len(pass2_inserts)} (flagged contracts not in XLSX)")


if __name__ == "__main__":
    main()
