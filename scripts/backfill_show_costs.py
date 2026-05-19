#!/usr/bin/env python3
"""
Backfill shows.total_cost from Lori's per-show Summary + Marketing + Travel
tabs.

For each XLSX:
  total_cost = sum(Marketing entries) + sum(Travel per-rep totals)
             + Summary's "Venue Costs" + "Setup Costs" + "Other Costs"

Matches yesterday's analytics migration 072 definition:
  "booth fees + travel + payroll + other"

Excludes COGS (Net Spa Costs, Cost of Delivery, Total Lump Freight Cost) —
those are per-unit costs that live on contracts already.

Usage:
    python3 scripts/backfill_show_costs.py [--dry-run] <xlsx-folder>
"""

import argparse
import os
import re
import sys

try:
    import openpyxl
    from supabase import create_client
except ImportError as e:
    print(f"ERROR: {e}", file=sys.stderr)
    sys.exit(1)

SHOW_FILE_RE = re.compile(r"^(\d{2})-(\d{2})-(\d{2})\s+(.+?)\s+(?:Boat\s+|Tent\s+)?Sales\.xlsx$")

OVERHEAD_SUMMARY_LABELS = {
    "Venue Costs": "venue",
    "Setup Costs": "setup",
    "Other Costs": "other",
}


def parse_filename(fn):
    m = SHOW_FILE_RE.match(fn)
    if not m:
        return None, None
    return f"20{m.group(3)}-{m.group(1)}-{m.group(2)}", m.group(4).split(",")[0].strip()


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def parse_costs(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    breakdown = {"marketing": 0.0, "travel": 0.0, "venue": 0.0, "setup": 0.0, "other": 0.0}

    # Marketing: col B summed for rows 5+
    if "Marketing" in wb.sheetnames:
        ws = wb["Marketing"]
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row and len(row) > 1:
                breakdown["marketing"] += num(row[1])

    # Travel: cols B-F summed for rows 4+ (Air, Hotel, Car Rental, Cab/Uber/Parking, Miles)
    if "Travel" in wb.sheetnames:
        ws = wb["Travel"]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[0]:
                continue
            for v in row[1:6]:
                breakdown["travel"] += num(v)

    # Summary tab: scan for labeled rows
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        for row_idx in range(1, min(60, ws.max_row + 1)):
            label = ws.cell(row=row_idx, column=1).value
            val = ws.cell(row=row_idx, column=2).value
            if isinstance(label, str):
                for label_text, bucket in OVERHEAD_SUMMARY_LABELS.items():
                    if label.strip() == label_text and isinstance(val, (int, float)):
                        breakdown[bucket] = float(val)

    return breakdown


def main():
    p = argparse.ArgumentParser()
    p.add_argument("folder", help="XLSX folder")
    p.add_argument("--dry-run", action="store_true", help="Print projected updates, no DB writes")
    args = p.parse_args()

    sb = create_client(
        os.environ["NEXT_PUBLIC_SUPABASE_URL"],
        os.environ["SUPABASE_SERVICE_ROLE_KEY"],
    )

    # Build lookup of backfilled shows by (start_date, city)
    shows = sb.table("shows").select("id, name, start_date, city, total_cost").eq("active", False).execute().data
    show_by_key = {(s["start_date"], (s["city"] or "").lower()): s for s in shows}

    file_paths = []
    for root, _, files in os.walk(args.folder):
        for fn in files:
            if fn.endswith(".xlsx") and not fn.startswith("~"):
                file_paths.append(os.path.join(root, fn))
    file_paths.sort(key=lambda p: os.path.basename(p))

    updated = 0
    skipped = 0
    grand_total = 0.0

    print(f"\n{'='*100}")
    print(f"{'File':<50}{'Marketing':>10}{'Travel':>10}{'Venue':>10}{'Setup':>10}{'Other':>10}{'TOTAL':>12}")
    print("=" * 100)

    for path in file_paths:
        fn = os.path.basename(path)
        iso, city = parse_filename(fn)
        if not iso:
            continue
        key = (iso, (city or "").lower())
        s = show_by_key.get(key)
        if not s:
            print(f"  {fn}: no matching show (skipped)")
            skipped += 1
            continue

        try:
            br = parse_costs(path)
        except Exception as e:
            print(f"  {fn}: parse error {e}")
            skipped += 1
            continue

        total = sum(br.values())
        grand_total += total
        short = fn[:48]
        print(
            f"{short:<50}{br['marketing']:>10,.0f}{br['travel']:>10,.0f}"
            f"{br['venue']:>10,.0f}{br['setup']:>10,.0f}{br['other']:>10,.0f}{total:>12,.0f}"
        )

        if not args.dry_run:
            sb.table("shows").update({"total_cost": round(total, 2)}).eq("id", s["id"]).execute()
        updated += 1

    print("=" * 100)
    mode = "DRY-RUN — no DB writes" if args.dry_run else f"UPDATED {updated} shows in DB"
    print(f"\n{mode}")
    print(f"Grand total historical overhead: ${grand_total:,.2f}")
    print(f"Files processed: {updated}, skipped: {skipped}")


if __name__ == "__main__":
    main()
