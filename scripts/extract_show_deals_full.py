#!/usr/bin/env python3
"""
Extract every INPUT field (not formula fields) from one Lori XLSX into JSON.
Used to round-trip-test the show-sales template + export pipeline:
extract from source → inject via TS → diff output against source.

Usage:
    python3 scripts/extract_show_deals_full.py <source.xlsx> <output.json>
"""

import json
import os
import sys
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed", file=sys.stderr)
    sys.exit(1)

# Map of (column letter, JSON field name) for INPUT fields only.
# Formula columns are excluded — the template will recompute them.
INPUT_COLUMNS = [
    ("A", "day_of_week"),
    ("B", "status"),
    ("D", "last_name"),
    ("E", "first_name"),
    ("F", "address"),
    ("G", "city"),
    ("H", "state"),
    ("I", "zip"),
    ("J", "salesman_1"),
    ("K", "salesman_2"),
    ("L", "salesman_3"),
    ("M", "salesman_4"),
    ("N", "model"),
    ("O", "color"),
    ("P", "color_cost"),
    ("Q", "cabinet"),
    ("R", "cabinet_cost"),
    ("S", "serial_number"),
    ("T", "masterpur"),
    ("U", "masterpur_cost"),
    ("V", "floor_system"),
    ("W", "floor_system_cost"),
    ("X", "other_options_1"),
    ("Y", "other_options_1_cost"),
    ("Z", "other_options_2"),
    ("AA", "other_options_2_cost"),
    ("AB", "other_spa_costs"),
    ("AC", "step"),
    ("AG", "freight_cost"),
    ("AH", "delivery_cost"),
    ("AI", "crane_cost"),
    ("AJ", "removal_cost"),
    ("AK", "cover_lift_type"),
    ("AL", "cover_lift_count"),
    ("AO", "sale_price"),
    ("AP", "delivery_cost_charged"),
    ("AQ", "cancelled_deal_sale_amount"),
    ("AR", "sales_tax_rate"),
    ("AT", "override_reason"),
    ("AU", "commission_rate"),
    ("AW", "spiff_reason"),
    ("AX", "spiff_amount"),
    ("AY", "spiff_payable"),
    ("BV", "cash_deposit"),
    ("BW", "check_deposit"),
    ("BX", "debit_deposit"),
    ("BY", "visa_deposit"),
    ("BZ", "mastercard_deposit"),
    ("CA", "discover_deposit"),
    ("CB", "amex_deposit"),
    ("CC", "finance_deposit"),
    ("CE", "financed_amount"),
    ("CF", "plan_number"),
    ("CG", "financing_cost"),
    ("CH", "approx_delivery_date"),
    ("CI", "marketing_feedback"),
    ("CJ", "comments"),
]


def cell_to_python(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def col_to_idx(letter):
    n = 0
    for c in letter:
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n


def main():
    if len(sys.argv) < 3:
        print(__doc__, file=sys.stderr)
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2]

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Sales"]

    deals = []
    first_data_row = 4
    for row_idx in range(first_data_row, ws.max_row + 1):
        deal = {}
        # Skip if status is empty (blank row)
        status_val = ws.cell(row=row_idx, column=col_to_idx("B")).value
        if not status_val:
            continue
        for letter, field in INPUT_COLUMNS:
            v = ws.cell(row=row_idx, column=col_to_idx(letter)).value
            v = cell_to_python(v)
            if v is not None and v != "":
                deal[field] = v
        if deal:
            deals.append(deal)

    # Pull show config from Variables tab
    vars_ws = wb["Variables"] if "Variables" in wb.sheetnames else None
    show_config = {"salesman_roster": []}
    if vars_ws:
        # Column B has labels ("Show Name", "Location", "Date", "Date of Last Day")
        # Column C has the actual values
        show_config["show_name"] = vars_ws.cell(row=2, column=3).value
        show_config["location"] = vars_ws.cell(row=3, column=3).value
        show_config["date_range"] = vars_ws.cell(row=4, column=3).value
        last_day = vars_ws.cell(row=5, column=3).value
        show_config["date_of_last_day"] = cell_to_python(last_day)
        # Salesman roster from col A rows 2-21
        for r in range(2, 22):
            v = vars_ws.cell(row=r, column=1).value
            if v:
                show_config["salesman_roster"].append(str(v))

    payload = {
        "source_file": os.path.basename(src),
        "show": show_config,
        "deals": deals,
        "deal_count": len(deals),
    }
    with open(out, "w") as fh:
        json.dump(payload, fh, indent=2, default=str)
    print(
        f"Extracted {len(deals)} deals + show config → {out}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
