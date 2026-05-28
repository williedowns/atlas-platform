#!/usr/bin/env python3
"""
sync_tax_rates.py — Pull sales tax rates from free state DOR feeds into Supabase.

DRAFT — review parsers per state before running against prod. The exact column
positions for TX EDI and OK COPO files are documented by the state but worth
verifying against the latest quarterly file.

Usage:
    python scripts/sync_tax_rates.py tx        # Texas only
    python scripts/sync_tax_rates.py ks        # Kansas only
    python scripts/sync_tax_rates.py ok        # Oklahoma only
    python scripts/sync_tax_rates.py la        # Louisiana (prints manual reminder)
    python scripts/sync_tax_rates.py all       # All four

Env required:
    SUPABASE_URL                  (e.g. https://xxx.supabase.co)
    SUPABASE_SERVICE_ROLE_KEY     (service role — bypasses RLS for write)

Cadence: run quarterly. State files update Jan 1, Apr 1, Jul 1, Oct 1.

Free state sources (verified 2026-05-27):
    TX:  https://comptroller.texas.gov/taxes/file-pay/edi/sales-tax-rates.php
    KS:  https://www.ksrevenue.gov/salesratechanges.html  (Pub 1700 .xlsx)
    OK:  https://oktax.csa.ou.edu/oktax  (vendor-integration files)
    LA:  Manual — no clean state-wide download. Verify per parish at:
         https://parishe-file.revenue.louisiana.gov/lookup/lookup.aspx
"""

from __future__ import annotations

import argparse
import os
import sys
import datetime as dt
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import requests
from openpyxl import load_workbook
from supabase import create_client, Client


# ─── Config ────────────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

CACHE_DIR = Path(__file__).parent / ".tax_rate_cache"
CACHE_DIR.mkdir(exist_ok=True)


# ─── Common types ──────────────────────────────────────────────────────────
@dataclass
class RawRate:
    state: str
    jurisdiction_code: str
    jurisdiction_name: str
    jurisdiction_type: str  # state | county | parish | city | transit | special | combined
    rate: float             # 0..0.20
    effective_date: str     # ISO YYYY-MM-DD
    source: str             # TX_EDI | KS_PUB1700 | OK_COPO | LA_MANUAL
    source_file: str


@dataclass
class ZipRate:
    zip: str
    state: str
    combined_rate: float
    state_rate: float
    local_rate: float
    jurisdictions: list[dict]  # [{name,type,rate}, ...]
    effective_date: str
    source: str


def _client() -> Client:
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        sys.exit("ERROR: set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY env vars")
    return create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)


def _current_quarter_start(today: dt.date | None = None) -> dt.date:
    today = today or dt.date.today()
    month = ((today.month - 1) // 3) * 3 + 1
    return dt.date(today.year, month, 1)


def _download(url: str, filename: str) -> Path:
    """Download to local cache. Re-download if file is missing or > 6 days old."""
    dest = CACHE_DIR / filename
    if dest.exists() and (dt.datetime.now() - dt.datetime.fromtimestamp(dest.stat().st_mtime)).days < 6:
        return dest
    print(f"  ↓ downloading {url}")
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    dest.write_bytes(resp.content)
    return dest


def _upsert_raw(client: Client, rows: Iterable[RawRate]) -> int:
    payload = [
        {
            "state": r.state,
            "jurisdiction_code": r.jurisdiction_code,
            "jurisdiction_name": r.jurisdiction_name,
            "jurisdiction_type": r.jurisdiction_type,
            "rate": round(r.rate, 5),
            "effective_date": r.effective_date,
            "source": r.source,
            "source_file": r.source_file,
        }
        for r in rows
    ]
    if not payload:
        return 0
    # Chunked upsert to avoid request-size limits.
    n = 0
    for i in range(0, len(payload), 500):
        chunk = payload[i : i + 500]
        client.table("tax_rates_raw").upsert(
            chunk, on_conflict="state,jurisdiction_code,effective_date"
        ).execute()
        n += len(chunk)
    return n


def _upsert_by_zip(client: Client, rows: Iterable[ZipRate]) -> int:
    payload = [
        {
            "zip": r.zip,
            "state": r.state,
            "combined_rate": round(r.combined_rate, 5),
            "state_rate": round(r.state_rate, 5),
            "local_rate": round(r.local_rate, 5),
            "jurisdictions": r.jurisdictions,
            "effective_date": r.effective_date,
            "source": r.source,
        }
        for r in rows
    ]
    if not payload:
        return 0
    n = 0
    for i in range(0, len(payload), 500):
        chunk = payload[i : i + 500]
        client.table("tax_rates_by_zip").upsert(
            chunk, on_conflict="zip,effective_date"
        ).execute()
        n += len(chunk)
    return n


# ─── TEXAS ────────────────────────────────────────────────────────────────
# Source: TX Comptroller "latest" rate file (overwritten each quarter).
#
# REAL FORMAT (validated 2026-05-27 against Q2 2026 file):
#   - Tab-separated values (NOT fixed-width)
#   - Lines 1-17 are header/period definitions; skip
#   - Lines 18+ are data: 12 tab fields per row, one row per named "place"
#   - Field layout per data row:
#       f0: place_name (city or community)
#       f1: city_ta_code      f2: city_rate       (or "n/a","0")
#       f3: county_name       f4: county_ta_code  f5: county_rate
#       f6: mta_name          f7: mta_ta_code     f8: mta_rate
#       f9: spd_name          f10: spd_ta_code    f11: spd_rate
#   - State 6.25% baseline is in line 1 header field 4 — added separately.
#   - A single place can appear MULTIPLE times if it spans multiple counties
#     (e.g. Dallas appears under both Collin and Dallas counties).
#
# IMPORTANT: This file is JURISDICTION-keyed, NOT ZIP-keyed. To resolve an
# arbitrary TX address to a rate, you also need the separate TX ZIP→TA-code
# lookup file (Phase A.2 of the build plan — URL TBD).
def sync_tx(client: Client, effective: dt.date) -> tuple[int, int]:
    filename = "TX_taxrates.txt"
    url = "https://comptroller.texas.gov/data/edi/sales-tax/taxrates.txt"
    try:
        local = _download(url, filename)
    except requests.HTTPError:
        print(f"  ⚠ TX file not found at {url} — confirm on the EDI page")
        raise

    raw_rows: list[RawRate] = []
    zip_rows: list[ZipRate] = []  # remains empty until Phase A.2 ZIP→TA-code join

    seen_jurisdictions: set[tuple[str, str]] = set()  # dedupe (code, type)

    def _emit_jurisdiction(name: str, code: str, rate_str: str, jtype: str) -> None:
        """Add one jurisdiction-level row to raw_rows if it's real and not seen."""
        code = (code or "").strip()
        name = (name or "").strip()
        if not code or code.lower() == "n/a":
            return
        try:
            rate = float((rate_str or "0").strip())
        except ValueError:
            return
        if rate <= 0:
            return
        key = (code, jtype)
        if key in seen_jurisdictions:
            return
        seen_jurisdictions.add(key)
        raw_rows.append(
            RawRate(
                state="TX",
                jurisdiction_code=code,
                jurisdiction_name=name or code,
                jurisdiction_type=jtype,
                rate=rate,
                effective_date=effective.isoformat(),
                source="TX_EDI",
                source_file=filename,
            )
        )

    with open(local, "r", encoding="utf-8", errors="replace") as f:
        for line_no, raw_line in enumerate(f, start=1):
            line = raw_line.rstrip("\r\n")
            if not line:
                continue
            fields = line.split("\t")
            # Real data rows have exactly 12 fields with a place name in f0.
            if len(fields) != 12:
                continue
            place = fields[0].strip()
            if not place:
                continue
            # Header rows in the first ~17 lines have period codes like "20262",
            # "202601" in f0 — skip them by checking that f0 isn't purely digits.
            if place.isdigit():
                continue
            # Emit each non-zero, non-n/a jurisdiction once
            _emit_jurisdiction(place, fields[1], fields[2], "city")
            _emit_jurisdiction(fields[3], fields[4], fields[5], "county")
            _emit_jurisdiction(fields[6], fields[7], fields[8], "transit")
            _emit_jurisdiction(fields[9], fields[10], fields[11], "special")

    # State 6.25% baseline
    raw_rows.append(
        RawRate(
            state="TX",
            jurisdiction_code="TX_STATE",
            jurisdiction_name="Texas state",
            jurisdiction_type="state",
            rate=0.0625,
            effective_date=effective.isoformat(),
            source="TX_EDI",
            source_file=filename,
        )
    )

    # NOTE: The EDI file is jurisdiction-keyed, not ZIP-keyed. To compute
    # tax_rates_by_zip we need a ZIP→jurisdiction mapping. The Comptroller
    # publishes a separate ZIP→TA Code lookup file (also linked from the EDI
    # page). Loading and joining that file is a TODO before this fully works
    # for ZIP-driven lookup. For now we write only tax_rates_raw and the
    # OTD calculator will fall back to tax_show_locations for venues.
    raw_count = _upsert_raw(client, raw_rows)
    zip_count = _upsert_by_zip(client, zip_rows)
    return raw_count, zip_count


# ─── KANSAS ───────────────────────────────────────────────────────────────
# Source: KDOR Publication 1700 (Excel .xlsx), updated quarterly.
# https://www.ksrevenue.gov/salesratechanges.html
def sync_ks(client: Client, effective: dt.date) -> tuple[int, int]:
    quarter_q = ((effective.month - 1) // 3) + 1
    filename = f"KS_Pub1700_{effective.year}Q{quarter_q}.xlsx"
    # URL pattern — confirm against the KDOR rate changes page for current quarter.
    url = (
        "https://www.ksrevenue.gov/forms-btpub-pub1700-"
        f"{effective.year}q{quarter_q}.xlsx"
    )
    try:
        local = _download(url, filename)
    except requests.HTTPError:
        print(f"  ⚠ KS file not at expected URL — check the ksrevenue.gov page")
        raise

    wb = load_workbook(local, data_only=True, read_only=True)
    ws = wb.active

    raw_rows: list[RawRate] = []
    zip_rows: list[ZipRate] = []

    # Pub 1700 has multiple sheets; the rate-by-jurisdiction sheet typically
    # has columns: Jurisdiction Code, City/County, State Rate, Local Rate,
    # Combined Rate, ZIP (where applicable).
    headers: dict[str, int] = {}
    for row in ws.iter_rows(values_only=True):
        if not headers:
            # First non-empty header row
            if row and any(isinstance(c, str) and "Rate" in c for c in row if c):
                headers = {str(c).strip().lower(): i for i, c in enumerate(row) if c}
            continue
        if not row or not row[0]:
            continue
        # Map common column names defensively — Pub 1700 layout has shifted
        # over the years; defensive lookup avoids brittle index access.
        def col(name: str):
            for k, idx in headers.items():
                if name.lower() in k:
                    return row[idx]
            return None

        code = col("code") or col("jurisdiction")
        name = col("city") or col("county") or col("name")
        combined = col("combined") or col("total")
        state_r = col("state rate") or 0.065
        local_r = col("local") or 0
        zip_v = col("zip")

        try:
            combined_f = float(combined) / 100 if combined and float(combined) > 1 else float(combined or 0)
            state_f = float(state_r) / 100 if state_r and float(state_r) > 1 else float(state_r or 0.065)
            local_f = float(local_r) / 100 if local_r and float(local_r) > 1 else float(local_r or 0)
        except (TypeError, ValueError):
            continue

        if combined_f <= 0:
            continue

        raw_rows.append(
            RawRate(
                state="KS",
                jurisdiction_code=str(code or "")[:32],
                jurisdiction_name=str(name or code or ""),
                jurisdiction_type="combined",
                rate=combined_f,
                effective_date=effective.isoformat(),
                source="KS_PUB1700",
                source_file=filename,
            )
        )

        if zip_v and str(zip_v).strip().isdigit() and len(str(zip_v).strip()) == 5:
            zip_rows.append(
                ZipRate(
                    zip=str(zip_v).strip(),
                    state="KS",
                    combined_rate=combined_f,
                    state_rate=state_f,
                    local_rate=local_f,
                    jurisdictions=[
                        {"name": "Kansas state", "type": "state", "rate": state_f},
                        {"name": str(name or ""), "type": "combined", "rate": local_f},
                    ],
                    effective_date=effective.isoformat(),
                    source="KS_PUB1700",
                )
            )

    return _upsert_raw(client, raw_rows), _upsert_by_zip(client, zip_rows)


# ─── OKLAHOMA ─────────────────────────────────────────────────────────────
# Source: OK Tax Commission COPO files (Center for Spatial Analysis / OU).
# https://oktax.csa.ou.edu/oktax
#
# The OTC publishes a CSV/TSV vendor-integration file each quarter listing
# COPO codes (county-city-other-political-subdivision) with combined rates.
def sync_ok(client: Client, effective: dt.date) -> tuple[int, int]:
    quarter_q = ((effective.month - 1) // 3) + 1
    filename = f"OK_COPO_{effective.year}Q{quarter_q}.csv"
    # URL pattern — confirm against oktax.csa.ou.edu for current quarter.
    url = (
        "https://oktax.csa.ou.edu/oktax/downloads/"
        f"copo_{effective.year}q{quarter_q}.csv"
    )
    try:
        local = _download(url, filename)
    except requests.HTTPError:
        print(f"  ⚠ OK COPO file not at expected URL — check oktax.csa.ou.edu")
        raise

    import csv
    raw_rows: list[RawRate] = []
    zip_rows: list[ZipRate] = []

    with open(local, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Defensive column lookup — exact headers vary.
            code = (row.get("COPO") or row.get("Code") or "").strip()
            name = (row.get("City") or row.get("Jurisdiction") or row.get("Name") or "").strip()
            combined_str = (row.get("Combined Rate") or row.get("Rate") or "").strip().rstrip("%")
            zip_v = (row.get("ZIP") or row.get("Zip") or "").strip()
            try:
                combined_f = float(combined_str)
                if combined_f > 1:  # already a percent
                    combined_f /= 100
            except ValueError:
                continue
            if combined_f <= 0:
                continue
            state_f = 0.045
            local_f = max(combined_f - state_f, 0)
            raw_rows.append(
                RawRate(
                    state="OK",
                    jurisdiction_code=code,
                    jurisdiction_name=name or code,
                    jurisdiction_type="combined",
                    rate=combined_f,
                    effective_date=effective.isoformat(),
                    source="OK_COPO",
                    source_file=filename,
                )
            )
            if zip_v and zip_v.isdigit() and len(zip_v) == 5:
                zip_rows.append(
                    ZipRate(
                        zip=zip_v,
                        state="OK",
                        combined_rate=combined_f,
                        state_rate=state_f,
                        local_rate=local_f,
                        jurisdictions=[
                            {"name": "Oklahoma state", "type": "state", "rate": state_f},
                            {"name": name, "type": "combined", "rate": local_f},
                        ],
                        effective_date=effective.isoformat(),
                        source="OK_COPO",
                    )
                )

    return _upsert_raw(client, raw_rows), _upsert_by_zip(client, zip_rows)


# ─── LOUISIANA ────────────────────────────────────────────────────────────
# No clean state-wide rate file. Home-rule parishes maintain independent
# collection. Atlas's policy: verify each LA show's rate by phoning the
# destination parish's Sales Tax Department 2 weeks before the show, then
# insert a row into tax_show_locations.
def sync_la(client: Client, effective: dt.date) -> tuple[int, int]:
    print("  ── Louisiana ──")
    print("  LA does not provide a downloadable state-wide rate file.")
    print("  Home-rule parishes (East Baton Rouge, Jefferson, Orleans, et al.)")
    print("  administer local sales tax independently. Rates can change with")
    print("  short notice.")
    print()
    print("  ACTION for each LA show:")
    print("    1. Identify the destination parish.")
    print("    2. Call the parish Sales Tax Department 2+ weeks before show.")
    print("       Lookup index: https://lulstb.com/")
    print("       State portal: https://parishe-file.revenue.louisiana.gov/lookup/lookup.aspx")
    print("    3. Confirm: state rate (5% as of 2025), parish rate, city rate,")
    print("       any convention/tourism/economic-development districts.")
    print("    4. INSERT a row into public.tax_show_locations with the verified")
    print("       combined rate, jurisdictions JSON, verified_by, verified_at.")
    print()
    print("  No rows written to tax_rates_raw or tax_rates_by_zip for LA.")
    return 0, 0


# ─── CLI ──────────────────────────────────────────────────────────────────
def main() -> int:
    parser = argparse.ArgumentParser(description="Sync state sales tax rates into Supabase")
    parser.add_argument(
        "state",
        choices=["tx", "ks", "ok", "la", "all"],
        help="Which state(s) to sync",
    )
    parser.add_argument(
        "--effective",
        help="Effective date YYYY-MM-DD (default: current quarter start)",
    )
    args = parser.parse_args()

    if args.effective:
        effective = dt.date.fromisoformat(args.effective)
    else:
        effective = _current_quarter_start()
    print(f"Effective quarter start: {effective.isoformat()}")

    client = _client()
    targets = ["tx", "ks", "ok", "la"] if args.state == "all" else [args.state]
    totals: dict[str, tuple[int, int]] = {}
    for s in targets:
        print(f"\n── Syncing {s.upper()} ──")
        try:
            fn = {"tx": sync_tx, "ks": sync_ks, "ok": sync_ok, "la": sync_la}[s]
            raw_n, zip_n = fn(client, effective)
            totals[s] = (raw_n, zip_n)
            print(f"  ✓ {s.upper()}: {raw_n} raw rows, {zip_n} ZIP rows")
        except Exception as e:
            print(f"  ✗ {s.upper()} FAILED: {e}")
            totals[s] = (0, 0)

    print("\n── Summary ──")
    for s, (raw_n, zip_n) in totals.items():
        print(f"  {s.upper()}: raw={raw_n}  by_zip={zip_n}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
