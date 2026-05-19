#!/usr/bin/env python3
"""
Audit the historical backfill against the source XLSX files.
For each XLSX, compares deal count in the file vs contracts on the
linked show in Supabase. Surfaces mismatches.
"""

import json
import os
import re
import sys

try:
    import openpyxl
    from supabase import create_client
except ImportError as e:
    print(f"ERROR: {e}", file=sys.stderr)
    sys.exit(1)

SHOW_FILE_RE = re.compile(r"^(\d{2})-(\d{2})-(\d{2})\s+(.+?)\s+(?:Boat\s+|Tent\s+)?Sales\.xlsx$")

VALID_STATUSES = {"OK", "Cancelled", "Low Deposit", "Contingent", "Financing Pending"}


def parse_filename(fn):
    m = SHOW_FILE_RE.match(fn)
    if not m:
        return None, None
    iso = f"20{m.group(3)}-{m.group(1)}-{m.group(2)}"
    city = m.group(4).split(",")[0].strip()
    return iso, city


def count_xlsx_deals(path):
    """Return count of valid deal rows in the XLSX."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Sales" not in wb.sheetnames:
        return 0
    ws = wb["Sales"]
    count = 0
    for row_idx in range(4, ws.max_row + 1):
        status = ws.cell(row=row_idx, column=2).value  # column B
        if not status:
            continue
        status = str(status).strip()
        if status not in VALID_STATUSES:
            continue
        # Must have at least one salesman
        for c in (10, 11, 12, 13):  # J,K,L,M
            if ws.cell(row=row_idx, column=c).value:
                count += 1
                break
    return count


def main():
    if len(sys.argv) < 2:
        print("Usage: audit_backfill.py <xlsx-folder>", file=sys.stderr)
        sys.exit(1)
    folder = sys.argv[1]

    sb = create_client(
        os.environ["NEXT_PUBLIC_SUPABASE_URL"],
        os.environ["SUPABASE_SERVICE_ROLE_KEY"],
    )

    # Get all backfilled shows
    shows = sb.table("shows").select("id, name, start_date, city").ilike(
        "name", "%[backfill 2026-05-19]%"
    ).execute().data
    show_by_key = {(s["start_date"], (s["city"] or "").lower()): s for s in shows}

    file_paths = []
    for root, _, files in os.walk(folder):
        for fn in files:
            if fn.endswith(".xlsx") and not fn.startswith("~"):
                file_paths.append(os.path.join(root, fn))
    file_paths.sort(key=lambda p: os.path.basename(p))

    audit_results = []
    perfect = 0
    mismatches = []
    for path in file_paths:
        fn = os.path.basename(path)
        iso, city = parse_filename(fn)
        if not iso:
            continue
        xlsx_count = count_xlsx_deals(path)

        # Find matching show
        key = (iso, (city or "").lower())
        s = show_by_key.get(key)
        if not s:
            audit_results.append({
                "file": fn, "xlsx_deals": xlsx_count,
                "db_show_id": None, "db_contracts": 0, "status": "NO_SHOW",
            })
            mismatches.append((fn, xlsx_count, 0, "no show found"))
            continue

        db_count = sb.table("contracts").select("id", count="exact").eq(
            "show_id", s["id"]
        ).execute().count

        audit_results.append({
            "file": fn, "xlsx_deals": xlsx_count,
            "db_show_id": s["id"], "db_contracts": db_count,
            "status": "MATCH" if db_count == xlsx_count else "MISMATCH",
        })
        if db_count == xlsx_count:
            perfect += 1
        else:
            mismatches.append((fn, xlsx_count, db_count, "count mismatch"))

    # Write audit report
    with open("/tmp/backfill-audit.json", "w") as f:
        json.dump(audit_results, f, indent=2)

    print(f"\n=== AUDIT SUMMARY ===")
    print(f"  Files checked: {len(file_paths)}")
    print(f"  Perfect matches (xlsx=db): {perfect}")
    print(f"  Mismatches: {len(mismatches)}")
    print(f"\n  Total XLSX deals: {sum(r['xlsx_deals'] for r in audit_results)}")
    print(f"  Total DB contracts: {sum(r['db_contracts'] for r in audit_results)}")
    if mismatches:
        print(f"\n--- MISMATCHES ---")
        for fn, xl, db, reason in mismatches[:25]:
            print(f"  {fn}: xlsx={xl}, db={db} ({reason})")

    print(f"\nFull report: /tmp/backfill-audit.json")


if __name__ == "__main__":
    main()
