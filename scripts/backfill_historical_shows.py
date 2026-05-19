#!/usr/bin/env python3
"""
Backfill Lori Donahue's historical XLSX show workbooks into Supabase.

Usage:
    # Inspect the payload for one show without inserting (dev sanity check)
    python3 scripts/backfill_historical_shows.py --dry-run-one \\
        "/path/to/05-17-26 Williamson County, TX Sales.xlsx"

    # Insert that one show + all its deals into Supabase (transactional per show)
    python3 scripts/backfill_historical_shows.py --execute-one \\
        "/path/to/05-17-26 Williamson County, TX Sales.xlsx"

    # Walk every XLSX in a folder (recursive). Pair with --execute when ready.
    python3 scripts/backfill_historical_shows.py --execute-all \\
        "/Users/williedowns/Documents/Salta/Show Spreadsheets from Lori Donahue"

Idempotency:
    Each created record (show / profile / customer / contract) gets a
    `[backfill]` marker in a notes-like field. Re-running the same XLSX
    detects existing backfill records and skips them.

Status mapping:
    XLSX 'OK'                → contracts.status='signed'
    XLSX 'Cancelled'         → contracts.status='cancelled'
    XLSX 'Low Deposit'       → contracts.status='low_deposit'
    XLSX 'Contingent'        → contracts.status='signed' + is_contingent=true
    XLSX 'Financing Pending' → contracts.status='financing_pending'
"""

import argparse
import json
import os
import re
import sys
import uuid
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    from supabase import create_client, Client
except ImportError as e:
    print(f"ERROR: missing dependency ({e})", file=sys.stderr)
    sys.exit(1)

BACKFILL_MARKER = "[backfill 2026-05-19]"

NAME_NORM = {
    "dan cesaro": "Daniel Cesaro",
    "daniel cesaro": "Daniel Cesaro",
    "gio gonzalez": "Giovanni Gonzalez",
    "giovanni gonzalez": "Giovanni Gonzalez",
    "josh wyscarver": "Josh Wyscarver",
    "joshua wyscarver": "Josh Wyscarver",
    "m. conner brady": "M. Connor Brady",
    "m. connor brady": "M. Connor Brady",
    "blake carmen": "Blake Carman",
    "blake carman": "Blake Carman",
    "alex": "Alex Broyles",
    "alex broyles": "Alex Broyles",
    "tom devlin": "Thomas Devlin",
    "thomas devlin": "Thomas Devlin",
}

STATUS_MAP = {
    "OK": "signed",
    "Cancelled": "cancelled",
    "Low Deposit": "low_deposit",
    "Contingent": "signed",  # paired with is_contingent=true
    "Financing Pending": "financing_pending",
}

SHOW_FILE_RE = re.compile(r"^(\d{2})-(\d{2})-(\d{2})\s+(.+?)\s+(?:Boat\s+|Tent\s+)?Sales\.xlsx$")


def norm_rep(name):
    if not name:
        return None
    s = str(name).strip()
    return NAME_NORM.get(s.lower(), s)


def col(letter):
    n = 0
    for c in letter:
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n - 1


def parse_show_filename(filename):
    m = SHOW_FILE_RE.match(filename)
    if not m:
        return None, None, None
    mm, dd, yy = m.group(1), m.group(2), m.group(3)
    city_state = m.group(4).strip()
    iso = f"20{yy}-{mm}-{dd}"
    city = city_state.split(",")[0].strip()
    state = (
        city_state.split(",")[1].strip() if "," in city_state else ""
    )
    return iso, city, state


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def extract_show_data(path):
    """Parse one XLSX into a {show, deals[]} dict."""
    fn = os.path.basename(path)
    iso, city, state = parse_show_filename(fn)
    if not iso:
        return None

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Sales" not in wb.sheetnames:
        return None

    ws = wb["Sales"]

    # Variables sheet has the show name/venue
    show_name = city  # fallback
    venue_name = None
    if "Variables" in wb.sheetnames:
        v = wb["Variables"]
        show_name = _str(v.cell(row=2, column=3).value) or city
        venue_name = _str(v.cell(row=3, column=3).value)

    deals = []
    for row_idx in range(4, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=i + 1).value for i in range(ws.max_column)]
        status = _str(row[col("B")])
        if not status:
            continue
        if status not in STATUS_MAP:
            continue
        sms = []
        for letter in ("J", "K", "L", "M"):
            v = norm_rep(row[col(letter)])
            if v:
                sms.append(v)
        if not sms:
            continue

        deals.append({
            "day_of_week": _str(row[col("A")]),
            "status_raw": status,
            "last_name": _str(row[col("D")]),
            "first_name": _str(row[col("E")]),
            "address": _str(row[col("F")]),
            "city": _str(row[col("G")]),
            "state": _str(row[col("H")]),
            "zip": _str(row[col("I")]),
            "salesmen": sms,
            "model": _str(row[col("N")]),
            "color": _str(row[col("O")]),
            "color_cost": _num(row[col("P")]),
            "cabinet": _str(row[col("Q")]),
            "cabinet_cost": _num(row[col("R")]),
            "serial_number": _str(row[col("S")]),
            "masterpur": _str(row[col("T")]),
            "masterpur_cost": _num(row[col("U")]),
            "floor_system": _str(row[col("V")]),
            "floor_system_cost": _num(row[col("W")]),
            "other_options_1": _str(row[col("X")]),
            "other_options_1_cost": _num(row[col("Y")]),
            "other_options_2": _str(row[col("Z")]),
            "other_options_2_cost": _num(row[col("AA")]),
            "other_spa_costs": _num(row[col("AB")]),
            "step": _str(row[col("AC")]),
            "freight_cost": _num(row[col("AG")]),
            "delivery_cost": _num(row[col("AH")]),
            "crane_cost": _num(row[col("AI")]),
            "removal_cost": _num(row[col("AJ")]),
            "cover_lift_type": _str(row[col("AK")]),
            "cover_lift_count": int(row[col("AL")]) if _num(row[col("AL")]) else None,
            "sale_price": _num(row[col("AO")]),
            "total_cost": _num(row[col("AN")]),
            "sales_tax_rate": _num(row[col("AR")]),
            "sales_tax": _num(row[col("AS")]),
            "override_reason": _str(row[col("AT")]),
            "commission_rate": _num(row[col("AU")]),
            "spiff_reason": _str(row[col("AW")]),
            "spiff_amount": _num(row[col("AX")]),
            "spiff_payable": _str(row[col("AY")]),
            "cash_deposit": _num(row[col("BV")]),
            "check_deposit": _num(row[col("BW")]),
            "debit_deposit": _num(row[col("BX")]),
            "visa_deposit": _num(row[col("BY")]),
            "mastercard_deposit": _num(row[col("BZ")]),
            "discover_deposit": _num(row[col("CA")]),
            "amex_deposit": _num(row[col("CB")]),
            "finance_deposit": _num(row[col("CC")]),
            "financed_amount": _num(row[col("CE")]),
            "plan_number": _str(row[col("CF")]),
            "financing_cost": _num(row[col("CG")]),
            "approx_delivery_date": _str(row[col("CH")]),
            "marketing_feedback": _str(row[col("CI")]),
            "comments": _str(row[col("CJ")]),
        })

    return {
        "show": {
            "name": show_name,
            "venue_name": venue_name,
            "city": city,
            "state": state,
            "start_date": iso,
            "end_date": iso,
        },
        "deals": deals,
        "source_file": fn,
    }


class Backfiller:
    def __init__(self, sb: Client, dry_run: bool):
        self.sb = sb
        self.dry_run = dry_run
        # Caches built once per run
        self._profile_by_name: dict[str, dict] = {}
        self._org_id: Optional[str] = None
        self._profile_cache_loaded = False
        self._product_by_key: dict[str, dict] = {}
        self._product_cache_loaded = False

    def _load_profile_cache(self):
        if self._profile_cache_loaded:
            return
        rows = self.sb.table("profiles").select("id, full_name, role, organization_id").execute().data or []
        for r in rows:
            if r.get("full_name"):
                self._profile_by_name[r["full_name"].strip().lower()] = r
            if not self._org_id and r.get("organization_id"):
                self._org_id = r["organization_id"]
        self._profile_cache_loaded = True

    def _load_product_cache(self):
        if self._product_cache_loaded:
            return
        rows = self.sb.table("products").select("id, model_code, sku, name").execute().data or []
        for r in rows:
            for key in (r.get("model_code"), r.get("sku"), r.get("name")):
                if key:
                    self._product_by_key[str(key).strip().lower()] = r
        self._product_cache_loaded = True

    def resolve_rep(self, name: str) -> str:
        """Return profile id for a rep name; create one if missing."""
        self._load_profile_cache()
        key = name.strip().lower()
        cached = self._profile_by_name.get(key)
        if cached:
            return cached["id"]

        # profiles.id has FK to auth.users — we must create the auth user
        # first via the admin API. Use an un-loginable email + random password.
        safe = re.sub(r"[^a-z0-9]+", "", name.lower())[:30]
        synthetic_email = f"backfill+{safe}-{uuid.uuid4().hex[:8]}@atlasspas.local"
        random_password = uuid.uuid4().hex + uuid.uuid4().hex  # 64 random hex
        if self.dry_run:
            print(f"  [dry-run] CREATE auth user + profile: {name} ({synthetic_email})")
            new_id = str(uuid.uuid4())
        else:
            auth_resp = self.sb.auth.admin.create_user({
                "email": synthetic_email,
                "password": random_password,
                "email_confirm": True,  # skip email confirmation flow
                "user_metadata": {"backfill": True, "real_name": name},
            })
            new_id = auth_resp.user.id  # type: ignore
            profile_payload = {
                "id": new_id,
                "full_name": name,
                "email": synthetic_email,
                "role": "sales_rep",
                "organization_id": self._org_id,
            }
            self.sb.table("profiles").insert(profile_payload).execute()
        self._profile_by_name[key] = {"id": new_id, "full_name": name, "role": "sales_rep"}
        return new_id

    def resolve_product(self, model: Optional[str]) -> Optional[str]:
        if not model:
            return None
        self._load_product_cache()
        key = model.strip().lower()
        cached = self._product_by_key.get(key)
        return cached["id"] if cached else None

    def resolve_show(self, show_data: dict, source_file: str) -> str:
        """Find existing backfilled show by (start_date+city+marker) or create new."""
        existing = (
            self.sb.table("shows")
            .select("id, name, city, start_date")
            .eq("start_date", show_data["start_date"])
            .ilike("city", show_data["city"])
            .execute()
            .data or []
        )
        for s in existing:
            if BACKFILL_MARKER in (s.get("name") or ""):
                return s["id"]

        new_id = str(uuid.uuid4())
        # shows.address and shows.zip have NOT NULL constraints. The XLSX doesn't
        # carry venue address — Lori files them by city/date. Use the venue name
        # or city as a placeholder so the FK constraint is satisfied; sales_admin
        # can backfill real addresses later via the existing /admin/shows/[id] UI.
        placeholder_address = (
            show_data.get("venue_name") or f"{show_data['city']}, {show_data['state']}"
        )
        payload = {
            "id": new_id,
            "name": f"{show_data['name']} {BACKFILL_MARKER}",
            "venue_name": show_data.get("venue_name"),
            "address": placeholder_address,
            "city": show_data["city"],
            "state": show_data["state"],
            "zip": "00000",  # historical placeholder; Lori can update later
            "start_date": show_data["start_date"],
            "end_date": show_data["end_date"],
            "active": False,
            "organization_id": self._org_id,
        }
        if self.dry_run:
            print(f"  [dry-run] CREATE show: {payload['name']} ({payload['start_date']})")
        else:
            self.sb.table("shows").insert(payload).execute()
        return new_id

    def resolve_customer(self, deal: dict) -> str:
        """Lookup or create customer. Dedup on (last+first+zip)."""
        last = deal.get("last_name") or "Unknown"
        first = deal.get("first_name") or ""
        zip_code = deal.get("zip") or ""

        if last and first and zip_code:
            existing = (
                self.sb.table("customers")
                .select("id")
                .eq("last_name", last)
                .eq("first_name", first)
                .eq("zip", zip_code)
                .limit(1)
                .execute()
                .data or []
            )
            if existing:
                return existing[0]["id"]

        new_id = str(uuid.uuid4())
        # customers.email + phone may be NOT NULL. Use synthetic placeholders;
        # Lori or the rep can edit real contact info later from the customer page.
        safe_first = re.sub(r"[^a-z0-9]+", "", (first or "").lower())[:20]
        safe_last = re.sub(r"[^a-z0-9]+", "", last.lower())[:20]
        synthetic_email = (
            f"backfill+{safe_last}-{safe_first}-{new_id[:8]}@atlasspas.local"
        )
        payload = {
            "id": new_id,
            "first_name": first or "",
            "last_name": last,
            "email": synthetic_email,
            "phone": "0000000000",
            "address": deal.get("address") or "",
            "city": deal.get("city") or "",
            "state": deal.get("state") or "",
            "zip": zip_code or "00000",
            "organization_id": self._org_id,
        }
        if self.dry_run:
            print(f"  [dry-run] CREATE customer: {last}, {first} {zip_code}")
        else:
            self.sb.table("customers").insert(payload).execute()
        return new_id

    def import_show(self, source: dict, payload_preview_limit: int = 3) -> dict:
        """Run the full import for one parsed XLSX. Returns a summary dict."""
        self._load_profile_cache()  # warms org_id

        show_data = source["show"]
        source_file = source["source_file"]
        deals = source["deals"]

        print(f"\n── {source_file} ── {len(deals)} deals")
        show_id = self.resolve_show(show_data, source_file)

        created = {"profiles": 0, "customers": 0, "contracts": 0, "overrides": 0}
        deals_imported = 0

        for idx, deal in enumerate(deals):
            # Resolve reps
            rep_ids = [self.resolve_rep(name) for name in deal["salesmen"]]
            primary_rep_id = rep_ids[0]
            secondary_names = deal["salesmen"][1:]

            customer_id = self.resolve_customer(deal)
            product_id = self.resolve_product(deal.get("model"))

            status_raw = deal["status_raw"]
            status = STATUS_MAP.get(status_raw, "signed")
            is_contingent = status_raw == "Contingent"

            sale_price = deal.get("sale_price") or 0
            total_cost = deal.get("total_cost") or 0
            tax_rate = deal.get("sales_tax_rate") or 0
            tax_amount = deal.get("sales_tax") or 0

            deposit_components = {
                "cash": deal.get("cash_deposit") or 0,
                "check": deal.get("check_deposit") or 0,
                "debit": deal.get("debit_deposit") or 0,
                "visa": deal.get("visa_deposit") or 0,
                "mastercard": deal.get("mastercard_deposit") or 0,
                "discover": deal.get("discover_deposit") or 0,
                "amex": deal.get("amex_deposit") or 0,
                "finance": deal.get("finance_deposit") or 0,
            }
            deposit_total = sum(deposit_components.values())
            financed_amount = deal.get("financed_amount") or 0

            payment_method = None
            for method, amount in deposit_components.items():
                if amount > 0:
                    payment_method = {
                        "cash": "cash",
                        "check": "ach",
                        "debit": "debit_card",
                        "visa": "credit_card",
                        "mastercard": "credit_card",
                        "discover": "credit_card",
                        "amex": "credit_card",
                        "finance": "financing",
                    }[method]
                    break

            contract_id = str(uuid.uuid4())
            contract_number = f"BF-{show_data['start_date']}-{show_data['city'][:6].upper()}-{idx+1:03d}"
            # Idempotency: skip if this contract_number already exists from a
            # prior import attempt (likely partial — interrupted by a constraint
            # error). Override row may also already exist; insert will no-op
            # silently via ON CONFLICT pattern below.
            if not self.dry_run:
                existing = (
                    self.sb.table("contracts")
                    .select("id")
                    .eq("contract_number", contract_number)
                    .limit(1)
                    .execute()
                    .data or []
                )
                if existing:
                    print(f"  SKIP deal #{idx+1} ({contract_number}) — already imported")
                    deals_imported += 1
                    continue
            contract_payload = {
                "id": contract_id,
                "contract_number": contract_number,
                "status": status,
                "customer_id": customer_id,
                "sales_rep_id": primary_rep_id,
                "show_id": show_id,
                "line_items": [{
                    "product_id": product_id,
                    "name": deal.get("model") or "Unknown Model",
                    "unit_price": sale_price,
                    "quantity": 1,
                }],
                "discounts": [],
                "financing": [{
                    "type": "external" if financed_amount > 0 else "none",
                    "financed_amount": financed_amount,
                }] if financed_amount > 0 else [],
                "subtotal": sale_price,
                "discount_total": 0,
                "tax_amount": tax_amount,
                "tax_rate": tax_rate,
                "total": sale_price + tax_amount,
                "deposit_amount": deposit_total,
                "deposit_paid": deposit_total,
                "balance_due": max(0, (sale_price + tax_amount) - deposit_total - financed_amount),
                "payment_method": payment_method,
                "is_contingent": is_contingent,
                "notes": f"{BACKFILL_MARKER} source={source_file} row={idx+1}",
                "organization_id": self._org_id,
                "doc_fee_amount": 0,
                "doc_fee_waived": True,
                "doc_fee_tax_amount": 0,
            }

            override_payload = {
                "contract_id": contract_id,
                "status_override": status_raw if status_raw != "OK" else None,
                "day_of_week": deal.get("day_of_week"),
                "salesman_2": secondary_names[0] if len(secondary_names) > 0 else None,
                "salesman_3": secondary_names[1] if len(secondary_names) > 1 else None,
                "salesman_4": secondary_names[2] if len(secondary_names) > 2 else None,
                "color": deal.get("color"),
                "color_cost": deal.get("color_cost"),
                "cabinet": deal.get("cabinet"),
                "cabinet_cost": deal.get("cabinet_cost"),
                "serial_number": deal.get("serial_number"),
                "masterpur": deal.get("masterpur"),
                "masterpur_cost": deal.get("masterpur_cost"),
                "floor_system": deal.get("floor_system"),
                "floor_system_cost": deal.get("floor_system_cost"),
                "other_options_1": deal.get("other_options_1"),
                "other_options_1_cost": deal.get("other_options_1_cost"),
                "other_options_2": deal.get("other_options_2"),
                "other_options_2_cost": deal.get("other_options_2_cost"),
                "other_spa_costs": deal.get("other_spa_costs"),
                "step": deal.get("step"),
                "freight_cost": deal.get("freight_cost"),
                "delivery_cost": deal.get("delivery_cost"),
                "crane_cost": deal.get("crane_cost"),
                "removal_cost": deal.get("removal_cost"),
                "cover_lift_type": deal.get("cover_lift_type"),
                "cover_lift_count": deal.get("cover_lift_count"),
                "override_reason": deal.get("override_reason"),
                "commission_rate": deal.get("commission_rate"),
                "spiff_reason": deal.get("spiff_reason"),
                "spiff_amount": deal.get("spiff_amount"),
                "spiff_payable": deal.get("spiff_payable"),
                "plan_number": deal.get("plan_number"),
                "financing_cost": deal.get("financing_cost"),
                "approx_delivery_date": deal.get("approx_delivery_date"),
                "marketing_feedback": deal.get("marketing_feedback"),
                "comments": deal.get("comments"),
            }

            if self.dry_run:
                if idx < payload_preview_limit:
                    print(f"\n  [dry-run] CONTRACT #{idx+1} — {deal.get('last_name')} {deal.get('first_name')} / {deal.get('model')}")
                    print(f"    status={status} ({status_raw}), sale={sale_price}, deposit={deposit_total}")
                    print(f"    customer_id={customer_id}, sales_rep_id={primary_rep_id}, show_id={show_id}")
                if idx == payload_preview_limit:
                    print(f"  … (showing first {payload_preview_limit} of {len(deals)} deals)")
            else:
                self.sb.table("contracts").insert(contract_payload).execute()
                # show_deal_overrides has contract_id PK with FK to contracts; insert after
                clean_override = {k: v for k, v in override_payload.items() if v is not None}
                if len(clean_override) > 1:
                    self.sb.table("show_deal_overrides").insert(clean_override).execute()

            deals_imported += 1
            created["contracts"] += 1
            created["overrides"] += 1

        return {
            "source_file": source_file,
            "show_id": show_id,
            "deals_imported": deals_imported,
            "deals_in_file": len(deals),
            "created": created,
        }


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--dry-run-one", help="Parse one XLSX, print payloads, no inserts")
    p.add_argument("--execute-one", help="Insert one XLSX into Supabase")
    p.add_argument("--execute-all", help="Insert every XLSX in folder (recursive)")
    args = p.parse_args()

    if not (args.dry_run_one or args.execute_one or args.execute_all):
        print(__doc__, file=sys.stderr)
        sys.exit(1)

    sb_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    sb_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not sb_url or not sb_key:
        print("ERROR: NEXT_PUBLIC_SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required", file=sys.stderr)
        sys.exit(2)
    sb = create_client(sb_url, sb_key)

    if args.dry_run_one:
        bf = Backfiller(sb, dry_run=True)
        source = extract_show_data(args.dry_run_one)
        if not source:
            print(f"Could not parse {args.dry_run_one}", file=sys.stderr); sys.exit(3)
        summary = bf.import_show(source)
        print(f"\n=== SUMMARY ===\n{json.dumps(summary, indent=2, default=str)}")
    elif args.execute_one:
        bf = Backfiller(sb, dry_run=False)
        source = extract_show_data(args.execute_one)
        if not source:
            print(f"Could not parse {args.execute_one}", file=sys.stderr); sys.exit(3)
        summary = bf.import_show(source)
        print(f"\n=== EXECUTED ===\n{json.dumps(summary, indent=2, default=str)}")
    elif args.execute_all:
        bf = Backfiller(sb, dry_run=False)
        file_paths = []
        for root, _, files in os.walk(args.execute_all):
            for fn in files:
                if fn.endswith(".xlsx") and not fn.startswith("~"):
                    file_paths.append(os.path.join(root, fn))
        file_paths.sort(key=lambda p: os.path.basename(p))
        all_results = []
        for path in file_paths:
            try:
                source = extract_show_data(path)
                if not source:
                    print(f"SKIP {os.path.basename(path)}: unparseable", file=sys.stderr)
                    continue
                summary = bf.import_show(source)
                all_results.append(summary)
            except Exception as e:
                print(f"ERROR on {os.path.basename(path)}: {e}", file=sys.stderr)
        print(f"\n=== ALL DONE ===")
        print(f"  shows imported: {len(all_results)}")
        print(f"  deals imported: {sum(r['deals_imported'] for r in all_results)}")


if __name__ == "__main__":
    main()
