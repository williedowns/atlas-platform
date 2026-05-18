#!/usr/bin/env python3
"""
Parse expo XLSX files from Lori Donahue and emit a static JSON dataset
consumed by /shows/leaderboard.

Usage:
    python3 scripts/build_show_leaderboard_data.py \
        "/Users/williedowns/Documents/Salta/Show Spreadsheets from Lori Donahue"

Outputs: src/lib/show-leaderboard/data.json

The JSON is a flat list of deal records plus a list of unique shows.
The ranking module consumes this at request time.
"""

import json
import os
import re
import sys
from datetime import date

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

NAME_NORM = {
    "dan cesaro": "Daniel Cesaro",
    "daniel cesaro": "Daniel Cesaro",
    "gio gonzalez": "Giovanni Gonzalez",
    "giovanni gonzalez": "Giovanni Gonzalez",
    "josh wyscarver": "Josh Wyscarver",
    "joshua wyscarver": "Josh Wyscarver",
    "m. conner brady": "M. Connor Brady",
    "m. connor brady": "M. Connor Brady",
    "blake carmen": "Blake Carman",
    "blake carman": "Blake Carman",
    "alex": "Alex Broyles",
    "alex broyles": "Alex Broyles",
    "tom devlin": "Thomas Devlin",
    "thomas devlin": "Thomas Devlin",
}


def norm(name):
    if not name:
        return None
    s = str(name).strip()
    return NAME_NORM.get(s.lower(), s)


SHOW_FILE_RE = re.compile(r"^(\d{2})-(\d{2})-(\d{2})\s+(.+?)\s+Sales\.xlsx$")


def parse_show_id(filename):
    m = SHOW_FILE_RE.match(filename)
    if not m:
        return None, None, None
    mm, dd, yy = m.group(1), m.group(2), m.group(3)
    city = m.group(4).strip()
    iso = f"20{yy}-{mm}-{dd}"
    show_id = f"{iso}__{re.sub(r'[^A-Za-z0-9]+', '-', city).strip('-').lower()}"
    return show_id, iso, city


def parse_file(path):
    """Return (show_meta, deal_rows[]) for one XLSX file."""
    filename = os.path.basename(path)
    show_id, iso_date, city = parse_show_id(filename)
    if not show_id:
        print(f"  skip (cannot parse filename): {filename}", file=sys.stderr)
        return None, []
    year = int(iso_date.split("-")[0])

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    def idx(label):
        try:
            return headers.index(label)
        except ValueError:
            return None

    i_status = idx("Status")
    i_sm_cols = [i for i, h in enumerate(headers) if h == "Salesman"][:4]
    i_sale = idx("Sale Price")
    i_cost = idx("Total Cost")
    i_lift = idx("# of Cover Lifts")
    i_model = idx("Model")

    if i_status is None or not i_sm_cols or i_sale is None or i_cost is None:
        print(f"  skip (missing required columns): {filename}", file=sys.stderr)
        return None, []

    deals = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        status = row[i_status]
        if not status:
            continue
        status = str(status).strip()
        salesmen = [norm(row[i]) for i in i_sm_cols if i < len(row) and row[i]]
        salesmen = [s for s in salesmen if s]
        if not salesmen:
            continue
        sale = row[i_sale] if i_sale is not None and row[i_sale] is not None else 0
        cost = row[i_cost] if i_cost is not None and row[i_cost] is not None else 0
        lift_raw = row[i_lift] if i_lift is not None else None
        has_lift = lift_raw not in (None, 0, "0", "")
        model = row[i_model] if i_model is not None else None

        deals.append({
            "show_id": show_id,
            "year": year,
            "status": status,
            "salesmen": salesmen,
            "sale_price": float(sale) if sale else 0.0,
            "total_cost": float(cost) if cost else 0.0,
            "has_lift": bool(has_lift),
            "model": str(model).strip() if model else None,
        })

    show_meta = {"id": show_id, "date": iso_date, "city": city, "year": year}
    return show_meta, deals


def main():
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        sys.exit(1)
    folder = sys.argv[1]
    if not os.path.isdir(folder):
        print(f"ERROR: folder not found: {folder}", file=sys.stderr)
        sys.exit(2)

    here = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.normpath(os.path.join(here, "..", "src", "lib", "show-leaderboard"))
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "data.json")

    shows = []
    all_deals = []
    # Walk subdirectories (e.g. 2025/, 2026/) plus the top-level folder.
    file_paths = []
    for root, _dirs, fnames in os.walk(folder):
        for fn in fnames:
            if fn.endswith(".xlsx") and not fn.startswith("~"):
                file_paths.append(os.path.join(root, fn))
    file_paths.sort(key=lambda p: os.path.basename(p))
    print(f"Processing {len(file_paths)} files from {folder} (recursive)", file=sys.stderr)
    for full in file_paths:
        show_meta, deals = parse_file(full)
        if show_meta is None:
            continue
        shows.append(show_meta)
        all_deals.extend(deals)
        print(f"  ok  {os.path.basename(full)}  ({len(deals)} rows)", file=sys.stderr)

    payload = {
        "generated_at": date.today().isoformat(),
        "source_folder": folder,
        "shows": shows,
        "deals": all_deals,
    }
    with open(out_path, "w") as fh:
        json.dump(payload, fh, indent=2)
    print(f"\nWrote {len(shows)} shows, {len(all_deals)} deals → {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
