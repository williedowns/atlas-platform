#!/usr/bin/env python3
"""
Backfill contracts.cost from Lori's historical XLSX workbooks.

Each backfilled contract has notes like:
    [backfill 2026-05-19] source=03-08-25 Tyler, TX Sales.xlsx row=7

This script:
  1. Reads every contract with a backfill marker
  2. Parses source filename and row index out of notes
  3. Opens the matching XLSX in the "Show Spreadsheets from Lori Donahue" tree
  4. Reads column AN (Total Cost) at the parsed row
  5. Updates contracts.cost

Usage:
    # Inspect what would be updated, no DB writes
    python3 scripts/import_contract_costs.py --dry-run

    # Actually update Supabase
    python3 scripts/import_contract_costs.py --execute

    # Override XLSX root folder
    python3 scripts/import_contract_costs.py --dry-run \\
        --root "/Users/williedowns/Documents/Salta/Show Spreadsheets from Lori Donahue"

Environment:
    NEXT_PUBLIC_SUPABASE_URL          (from .env.local)
    SUPABASE_SERVICE_ROLE_KEY         (from .env.local)
"""

import argparse
import os
import re
import sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from supabase import create_client
except ImportError as e:
    print(f"ERROR: missing dependency ({e})", file=sys.stderr)
    sys.exit(1)


DEFAULT_XLSX_ROOT = "/Users/williedowns/Documents/Salta/Show Spreadsheets from Lori Donahue"
NOTE_RE = re.compile(r"source=(?P<file>.+?)\s+row=(?P<row>\d+)")


def col(letter: str) -> int:
    """A=0, B=1, ..., AA=26, AN=39"""
    n = 0
    for c in letter:
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n - 1


# Status filter that the original backfill applied — only deals with these
# statuses entered the deals[] list. Must match backfill_historical_shows.py
# exactly or our (file, idx) → cost mapping will be wrong.
STATUS_MAP_KEYS = {"OK", "Cancelled", "Low Deposit", "Contingent", "Financing Pending"}


def find_xlsx(root: Path, filename: str) -> "Path | None":
    """Locate filename anywhere under root (recursive)."""
    for p in root.rglob(filename):
        if p.is_file():
            return p
    return None


def extract_deal_costs(path: Path):
    """Re-extract deals from an XLSX using the SAME loop as
    backfill_historical_shows.py, returning a list of (idx+1, total_cost)
    keyed by 1-based position in the extracted deals[] list.

    The contracts.notes `row=N` value refers to this position, NOT to the
    XLSX row number. Header / totals / empty rows are skipped before
    enumeration, so we must mirror the skip logic precisely.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    out = {}
    deal_idx = 0
    cost_col_idx = col("AN")
    status_col_idx = col("B")
    salesman_cols = [col(L) for L in ("J", "K", "L", "M")]
    for row_idx in range(4, ws.max_row + 1):
        status_val = ws.cell(row=row_idx, column=status_col_idx + 1).value
        status = str(status_val).strip() if status_val is not None else ""
        if not status or status not in STATUS_MAP_KEYS:
            continue
        # Must have at least one salesman, per backfill logic
        has_salesman = False
        for c_idx in salesman_cols:
            v = ws.cell(row=row_idx, column=c_idx + 1).value
            if v and str(v).strip():
                has_salesman = True
                break
        if not has_salesman:
            continue
        deal_idx += 1
        cost_val = ws.cell(row=row_idx, column=cost_col_idx + 1).value
        out[deal_idx] = (cost_val, row_idx)
    wb.close()
    return out


def parse_notes(notes: "str | None") -> tuple["str | None", "int | None"]:
    if not notes:
        return None, None
    m = NOTE_RE.search(notes)
    if not m:
        return None, None
    try:
        return m.group("file"), int(m.group("row"))
    except (ValueError, TypeError):
        return None, None


def load_env():
    env_path = Path(__file__).parent.parent / ".env.local"
    env = {}
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            m = re.match(r"^([A-Z_]+)\s*=\s*(.*?)\s*$", line)
            if m:
                env[m.group(1)] = m.group(2).strip("\"'")
    return env


def fetch_all_backfilled(sb):
    """Fetch all contracts whose notes contain the backfill marker."""
    out = []
    page, size = 0, 1000
    while True:
        r = sb.table("contracts") \
            .select("id, contract_number, notes, total, cost") \
            .like("notes", "%[backfill%") \
            .range(page * size, (page + 1) * size - 1) \
            .execute()
        chunk = r.data or []
        out.extend(chunk)
        if len(chunk) < size:
            break
        page += 1
    return out


def to_amount(v) -> "float | None":
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace("$", "").replace(",", "").strip()
        return float(s) if s else None
    except Exception:
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="Parse + summarize, no DB writes")
    ap.add_argument("--execute", action="store_true", help="Actually update contracts.cost")
    ap.add_argument("--root", default=DEFAULT_XLSX_ROOT, help="Root folder containing the Lori XLSX workbooks")
    ap.add_argument("--limit", type=int, default=0, help="Process only this many contracts (debugging)")
    args = ap.parse_args()

    if not args.dry_run and not args.execute:
        print("ERROR: pass --dry-run or --execute", file=sys.stderr)
        sys.exit(2)

    env = load_env()
    url = env.get("NEXT_PUBLIC_SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: missing supabase env vars", file=sys.stderr)
        sys.exit(1)
    sb = create_client(url, key)

    root = Path(args.root)
    if not root.exists():
        print(f"ERROR: XLSX root not found: {root}", file=sys.stderr)
        sys.exit(1)

    print(f"  Fetching backfilled contracts from Supabase...")
    contracts = fetch_all_backfilled(sb)
    print(f"  Found {len(contracts)} backfilled contracts.")

    if args.limit:
        contracts = contracts[:args.limit]
        print(f"  (--limit {args.limit} applied, processing {len(contracts)})")

    # Cache file lookups + extracted deal maps per source file
    file_resolved_cache = {}
    deals_by_file = {}  # filename -> { deal_idx: (cost_val, xlsx_row) }

    updates = []
    skipped_no_marker = 0
    skipped_no_file = 0
    skipped_no_value = 0
    matched_zero_cost = 0
    skipped_already_set = 0
    skipped_index_oob = 0
    total_cost_sum = 0.0

    by_file_stats = defaultdict(lambda: {"updated": 0, "skipped": 0, "sum": 0.0})

    for c in contracts:
        if c.get("cost") is not None:
            skipped_already_set += 1
            continue

        src_file, deal_idx = parse_notes(c.get("notes"))
        if not src_file or not deal_idx:
            skipped_no_marker += 1
            continue

        # Resolve + extract deals from the XLSX (once per file)
        if src_file not in file_resolved_cache:
            file_resolved_cache[src_file] = find_xlsx(root, src_file)
        path = file_resolved_cache[src_file]
        if not path:
            skipped_no_file += 1
            by_file_stats[src_file]["skipped"] += 1
            continue

        if src_file not in deals_by_file:
            try:
                deals_by_file[src_file] = extract_deal_costs(path)
            except Exception as e:
                print(f"  [skip] cannot extract {path}: {e}", file=sys.stderr)
                deals_by_file[src_file] = {}
                skipped_no_file += 1
                continue

        deal_map = deals_by_file[src_file]
        if deal_idx not in deal_map:
            skipped_index_oob += 1
            continue

        cost_raw, xlsx_row = deal_map[deal_idx]
        cost_value = to_amount(cost_raw)

        if cost_value is None:
            skipped_no_value += 1
            continue

        if cost_value == 0:
            matched_zero_cost += 1
            continue

        updates.append({
            "id": c["id"],
            "cost": cost_value,
            "contract_number": c.get("contract_number"),
            "total": c.get("total"),
            "source_file": src_file,
            "deal_idx": deal_idx,
            "xlsx_row": xlsx_row,
        })
        total_cost_sum += cost_value
        by_file_stats[src_file]["updated"] += 1
        by_file_stats[src_file]["sum"] += cost_value

    # ── Summary ────────────────────────────────────────────────────────────
    print()
    print("─" * 70)
    print(f"Updates ready:        {len(updates):>5}")
    print(f"Skipped:")
    print(f"  no source marker:   {skipped_no_marker:>5}")
    print(f"  XLSX not found:     {skipped_no_file:>5}")
    print(f"  deal idx OOB:       {skipped_index_oob:>5}   (extracted deals list shorter than expected)")
    print(f"  no AN value:        {skipped_no_value:>5}")
    print(f"  AN = 0:             {matched_zero_cost:>5}")
    print(f"  cost already set:   {skipped_already_set:>5}")
    print(f"Total cost to import: ${total_cost_sum:,.2f}")
    print("─" * 70)

    # Sample + sanity-check distribution
    if updates:
        margins = []
        for u in updates:
            t = float(u["total"]) if u["total"] else 0
            if t > 0:
                margins.append((t - u["cost"]) / t * 100)
        if margins:
            margins.sort()
            mid = len(margins) // 2
            print(f"\nMargin distribution across {len(margins)} contracts:")
            print(f"  median:  {margins[mid]:>6.1f}%")
            print(f"  p10:     {margins[len(margins)//10]:>6.1f}%")
            print(f"  p90:     {margins[(len(margins)*9)//10]:>6.1f}%")
            print(f"  worst:   {margins[0]:>6.1f}%")
            print(f"  best:    {margins[-1]:>6.1f}%")
            negative_margin_count = sum(1 for m in margins if m < 0)
            print(f"  negative-margin contracts: {negative_margin_count}")

        print("\nSample (5 random):")
        import random
        for u in random.sample(updates, min(5, len(updates))):
            t = float(u["total"]) if u["total"] else 0
            margin_pct = ((t - u["cost"]) / t * 100) if t else 0
            print(f"  #{u['contract_number']:<28}  total=${t:>9,.2f}  cost=${u['cost']:>9,.2f}  margin={margin_pct:>6.1f}%   ({u['source_file']} deal#{u['deal_idx']}, xlsx_row={u['xlsx_row']})")

    if args.dry_run:
        print("\nDry run complete. Re-run with --execute to apply.")
        return

    # ── Execute ────────────────────────────────────────────────────────────
    print(f"\nWriting {len(updates)} cost updates to Supabase...")
    written = 0
    # supabase-py doesn't have batch update; loop with PATCH per row.
    # 2,118 updates over a few minutes is acceptable.
    for u in updates:
        try:
            sb.table("contracts").update({"cost": u["cost"]}).eq("id", u["id"]).execute()
            written += 1
            if written % 100 == 0:
                print(f"  ... {written}/{len(updates)}")
        except Exception as e:
            print(f"  [error] contract {u['contract_number']}: {e}", file=sys.stderr)

    print(f"\nWrote {written} cost values. {len(updates) - written} failed.")


if __name__ == "__main__":
    main()
